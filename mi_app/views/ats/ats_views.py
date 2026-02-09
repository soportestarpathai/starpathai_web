"""
Vistas para el producto ATS (Applicant Tracking System) de Star Path.
- Página de oferta / producto: "Ven y prueba nuestro ATS"
- Plataforma ATS: login, registro y dashboard para clientes.
- Dashboard: candidatos, habilidades (en detalle), cuenta/billing.
"""
from django.core.paginator import Paginator
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import TemplateView, View
from django.contrib.auth import login, logout, get_user_model
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import (
    PasswordChangeView as AuthPasswordChangeView,
    PasswordResetView as AuthPasswordResetView,
    PasswordResetDoneView as AuthPasswordResetDoneView,
    PasswordResetConfirmView as AuthPasswordResetConfirmView,
    PasswordResetCompleteView as AuthPasswordResetCompleteView,
)
from django.urls import reverse_lazy, reverse

from django.conf import settings
from django.core.cache import cache
from django.contrib import messages
from django.http import HttpResponseForbidden, HttpResponseRedirect, HttpResponse
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, FormView
from django.utils import timezone

from mi_app.views.ats.forms import (
    ATSRegisterForm,
    ATSLoginForm,
    ATSFormCreateEditForm,
    get_ats_form_field_formset,
    get_ats_form_criterion_formset,
    ATSEmailConfigForm,
    ATSProfileForm,
    ATSVacancyForm,
    CVAnalysisConfigForm,
)
from django.db.models import Count, Q, Sum  # Count for annotate, Q for filter
from datetime import timedelta

from mi_app.models import (
    ATSClient,
    Subscription,
    Candidate,
    Vacancy,
    CVAnalysisConfig,
    ATSForm,
    ATSFormField,
    ATSFormCriterion,
    ATSFormSubmission,
    ATSFormSubmissionFile,
    ATSCandidateCriterionResponse,
    ATSClientEmailConfig,
    ATSNotification,
    PlanChangeRequest,
    LLMUsageLog,
)
from mi_app.ats_plans import (
    get_all_plans,
    get_plan_capabilities_display,
    get_plan_candidates_limit,
    get_plan_vacancies_limit,
    subscription_can,
    subscription_can_add_candidate,
    subscription_can_add_vacancy,
    apply_plan_to_subscription,
)
from mi_app.ats_notifications import notify_ats_client, notify_support_plan_change, notify_support_account_deletion_request, send_email_to_candidate

User = get_user_model()


class StaffRequiredMixin(LoginRequiredMixin):
    """Solo usuarios con is_staff pueden acceder. Para el panel de administración ATS."""
    login_url = reverse_lazy("ats_plataforma")

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        if not request.user.is_staff:
            return HttpResponseForbidden("No tienes permiso para acceder a esta página.")
        return super().dispatch(request, *args, **kwargs)


class ATSProductoView(TemplateView):
    """Página que ofrece el producto ATS: Low Code, identificación de talento, formularios, OpenAI, match."""
    template_name = "ats/producto_ats.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["plans"] = get_all_plans()
        return context


class ATSPlataformaView(View):
    """Página de acceso: login y registro. Si ya está autenticado, redirige al dashboard."""
    template_name = "ats/plataforma.html"

    def get(self, request):
        if request.user.is_authenticated:
            return redirect("ats_dashboard")
        return render(request, self.template_name, {
            "login_form": ATSLoginForm(request),
            "register_form": ATSRegisterForm(),
        })

    def post(self, request):
        # El POST se maneja en las URLs específicas (login o register)
        return self.get(request)


class ATSRegisterView(View):
    """Registro de nuevo cliente ATS."""
    template_name = "ats/plataforma.html"
    success_url = reverse_lazy("ats_dashboard")

    def get(self, request):
        if request.user.is_authenticated:
            return redirect("ats_dashboard")
        return redirect("ats_plataforma")

    def post(self, request):
        if request.user.is_authenticated:
            return redirect("ats_dashboard")
        form = ATSRegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user, backend="django.contrib.auth.backends.ModelBackend")
            return redirect(self.success_url)
        return render(request, self.template_name, {
            "login_form": ATSLoginForm(request),
            "register_form": form,
            "active_tab": "register",
        })


class ATSLoginView(View):
    """Inicio de sesión ATS (correo + contraseña)."""
    template_name = "ats/plataforma.html"
    success_url = reverse_lazy("ats_dashboard")

    def get(self, request):
        if request.user.is_authenticated:
            if request.user.is_staff:
                return redirect("ats_admin_dashboard")
            return redirect("ats_dashboard")
        return redirect("ats_plataforma")

    def post(self, request):
        if request.user.is_authenticated:
            if request.user.is_staff:
                return redirect("ats_admin_dashboard")
            return redirect("ats_dashboard")
        form = ATSLoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user, backend="django.contrib.auth.backends.ModelBackend")
            if request.GET.get("next"):
                return redirect(request.GET["next"])
            if user.is_staff:
                return redirect("ats_admin_dashboard")
            return redirect(self.success_url)
        return render(request, self.template_name, {
            "login_form": form,
            "register_form": ATSRegisterForm(),
            "active_tab": "login",
        })


class ATSLogoutView(View):
    """Cerrar sesión y volver a la página de plataforma."""
    def get(self, request):
        logout(request)
        return redirect("ats_plataforma")

    def post(self, request):
        logout(request)
        return redirect("ats_plataforma")


def _get_or_create_subscription(user):
    """Crea suscripción FREE si el usuario no tiene (ej. clientes creados antes de este módulo)."""
    sub, _ = Subscription.objects.get_or_create(
        user=user,
        defaults={
            "plan": Subscription.PLAN_FREE,
            "cvs_limit": 3,
            "active": True,
        },
    )
    return sub


def user_can_process_cv(user):
    """
    Verifica si el usuario puede procesar un CV más (límite de plan).
    Usar antes de aceptar subida de CV / llamada a LLM; si falla, devolver 403 o mensaje claro.
    """
    sub = _get_or_create_subscription(user)
    return sub.can_process_cv


class ATSDashboardView(LoginRequiredMixin, TemplateView):
    """Panel del cliente ATS: KPIs, candidatos (filtro), gráfica, reclutamiento (vacantes), cuenta."""
    template_name = "ats/dashboard.html"
    login_url = reverse_lazy("ats_plataforma")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ats_client = getattr(self.request.user, "ats_client", None)
        context["ats_client"] = ats_client
        subscription = _get_or_create_subscription(self.request.user)
        context["subscription"] = subscription
        section = self.request.GET.get("section", "candidatos")
        context["dashboard_section"] = section
        context["ats_page"] = section
        # Planes con sus capacidades (vistas/procedimientos/actividades por plan)
        all_plans = get_all_plans()
        for p in all_plans:
            p["capabilities_display"] = get_plan_capabilities_display(p["id"])
        context["available_plans"] = all_plans
        context["subscription_plan_capabilities"] = get_plan_capabilities_display(subscription.plan)
        context["subscription_can_process_cv"] = subscription_can(subscription, "cvs_scan")
        context["subscription_can_export_candidates"] = subscription_can(subscription, "export_candidates")

        if ats_client:
            base_qs = Candidate.objects.filter(client=ats_client).select_related("vacancy")
            # KPIs (totales)
            context["kpi_total"] = base_qs.count()
            context["kpi_aptos"] = base_qs.filter(status=Candidate.STATUS_APTO).count()
            context["kpi_revision"] = base_qs.filter(status=Candidate.STATUS_REVISION).count()
            context["kpi_no_aptos"] = base_qs.filter(status=Candidate.STATUS_NO_APTO).count()
            context["kpi_cvs_used"] = subscription.cvs_used
            context["kpi_cvs_limit"] = subscription.cvs_limit
            context["kpi_candidates_limit"] = get_plan_candidates_limit(subscription.plan)
            context["kpi_vacancies_limit"] = get_plan_vacancies_limit(subscription.plan)
            context["kpi_vacancy_count"] = Vacancy.objects.filter(client=ats_client).count()
            # Datos para gráfica (por estado)
            context["chart_aptos"] = context["kpi_aptos"]
            context["chart_revision"] = context["kpi_revision"]
            context["chart_no_aptos"] = context["kpi_no_aptos"]
            # Filtros
            q = (self.request.GET.get("q") or "").strip()
            status_filter = self.request.GET.get("status", "")
            vacancy_id = self.request.GET.get("vacancy", "")
            qs = base_qs
            if q:
                qs = qs.filter(Q(name__icontains=q) | Q(email__icontains=q))
            if status_filter and status_filter in (Candidate.STATUS_APTO, Candidate.STATUS_REVISION, Candidate.STATUS_NO_APTO):
                qs = qs.filter(status=status_filter)
            if vacancy_id:
                try:
                    qs = qs.filter(vacancy_id=int(vacancy_id))
                except ValueError:
                    pass
            context["candidates"] = qs.prefetch_related("skill_evaluations")[:200]
            context["filter_q"] = q
            context["filter_status"] = status_filter
            context["filter_vacancy"] = vacancy_id
            # Vacantes (para filtro y sección Reclutamiento)
            context["vacancies"] = Vacancy.objects.filter(client=ats_client).annotate(
                candidates_count=Count("candidates")
            ).order_by("-created_at")
            # Envíos de formularios con vacante pero sin candidato (para botón "Crear candidatos desde envíos")
            context["pending_submissions_count"] = ATSFormSubmission.objects.filter(
                form__client=ats_client,
                form__vacancy__isnull=False,
                candidate__isnull=True,
            ).count()
        else:
            context["kpi_total"] = context["kpi_aptos"] = context["kpi_revision"] = context["kpi_no_aptos"] = 0
            context["kpi_cvs_used"] = subscription.cvs_used
            context["kpi_cvs_limit"] = subscription.cvs_limit
            context["kpi_candidates_limit"] = get_plan_candidates_limit(subscription.plan)
            context["kpi_vacancies_limit"] = get_plan_vacancies_limit(subscription.plan)
            context["kpi_vacancy_count"] = 0
            context["chart_aptos"] = context["chart_revision"] = context["chart_no_aptos"] = 0
            context["candidates"] = []
            context["filter_q"] = context["filter_status"] = context["filter_vacancy"] = ""
            context["vacancies"] = []
            context["pending_submissions_count"] = 0
            context["subscription_can_export_candidates"] = subscription_can(subscription, "export_candidates")
        return context


class ATSChangePlanView(LoginRequiredMixin, View):
    """
    El cliente solicita cambio de plan. No se aplica el cambio aquí: se envía correo
    a soporte para validar, gestionar pago externo y activar el plan desde el admin.
    """
    login_url = reverse_lazy("ats_plataforma")
    http_method_names = ["post"]

    def post(self, request):
        plan_id = (request.POST.get("plan") or "").strip().upper()
        if plan_id not in (Subscription.PLAN_FREE, Subscription.PLAN_PRO, Subscription.PLAN_ENTERPRISE):
            messages.warning(request, "Plan no válido.")
            return redirect(reverse("ats_dashboard") + "?section=cuenta")
        subscription = _get_or_create_subscription(request.user)
        old_plan_id = subscription.plan
        if old_plan_id == plan_id:
            messages.info(request, "Ya tienes ese plan seleccionado.")
            return redirect(reverse("ats_dashboard") + "?section=cuenta")
        ats_client = _get_client_or_403(request)
        # Enviar correo a soporte para validar, cobro externo y activar desde admin
        notify_support_plan_change(request.user, ats_client, old_plan_id, plan_id)
        if ats_client:
            PlanChangeRequest.objects.create(
                client=ats_client,
                from_plan=old_plan_id,
                to_plan=plan_id,
                status=PlanChangeRequest.STATUS_PENDING,
            )
        plan_name = dict(Subscription.PLAN_CHOICES).get(plan_id, plan_id)
        if ats_client:
            notify_ats_client(
                ats_client,
                ATSNotification.TYPE_PLAN,
                "Solicitud de cambio de plan",
                message=f"Solicitud enviada para pasar a {plan_name}. Soporte validará y te contactará; al gestionar el pago se activará tu nuevo plan.",
                link=reverse("ats_dashboard") + "?section=cuenta",
                request=request,
            )
        messages.success(
            request,
            "Solicitud enviada. Soporte validará el cambio y te contactará; una vez gestionado el pago, se activará tu nuevo plan desde administración.",
        )
        return redirect(reverse("ats_dashboard") + "?section=cuenta")


class ATSRequestAccountDeletionView(LoginRequiredMixin, View):
    """Solicitar baja de cuenta: envía correo a soporte y redirige a Mi cuenta."""
    login_url = reverse_lazy("ats_plataforma")

    def get(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("ats_dashboard")
        return render(request, "ats/request_account_deletion.html", {
            "ats_client": client,
            "ats_page": "cuenta",
        })

    def post(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("ats_dashboard")
        notify_support_account_deletion_request(client)
        messages.success(
            request,
            "Tu solicitud de baja ha sido enviada a soporte. Te contactaremos para confirmar.",
        )
        return redirect(reverse("ats_dashboard") + "?section=cuenta")


class ATSCandidateDetailView(LoginRequiredMixin, View):
    """Detalle de candidato: score, estado, explicación, habilidades y evaluación manual (Cumple/No cumple)."""
    template_name = "ats/candidate_detail.html"
    login_url = reverse_lazy("ats_plataforma")

    def _get_candidate_and_form(self, request, pk):
        ats_client = getattr(request.user, "ats_client", None)
        if not ats_client:
            return None, None
        candidate = get_object_or_404(
            Candidate.objects.prefetch_related(
                "skill_evaluations",
                "criterion_responses",
                "form_submissions__files",
                "form_submissions__form",
            ),
            pk=pk,
            client=ats_client,
        )
        ats_form = None
        first_sub = candidate.form_submissions.first()
        if first_sub:
            ats_form = first_sub.form
        return candidate, ats_form

    def get(self, request, pk):
        candidate, ats_form = self._get_candidate_and_form(request, pk)
        if not candidate:
            return render(request, self.template_name, {"candidate": None, "ats_page": "candidatos"})
        criteria = list(ats_form.criteria.all()) if ats_form else []
        response_map = {r.criterion_id: r.cumple for r in candidate.criterion_responses.filter(criterion__in=criteria)}
        evaluation_criteria_with_response = [(c, response_map.get(c.id, False)) for c in criteria]
        form_submission = candidate.form_submissions.first()
        subscription = _get_or_create_subscription(request.user)
        from django.conf import settings as django_settings
        cv_max = getattr(django_settings, "ATS_FORM_PUBLIC_MAX_FILE_SIZE", 10 * 1024 * 1024)
        context = {
            "candidate": candidate,
            "ats_client": getattr(request.user, "ats_client", None),
            "subscription": subscription,
            "ats_page": "candidatos",
            "form_submission": form_submission,
            "evaluation_criteria": criteria,
            "evaluation_criteria_with_response": evaluation_criteria_with_response,
            "subscription_can_process_cv": subscription_can(subscription, "cvs_scan"),
            "subscription_can_custom_email": subscription_can(subscription, "custom_email_message"),
            "kpi_cvs_used": subscription.cvs_used if subscription else 0,
            "kpi_cvs_limit": subscription.cvs_limit if subscription else 0,
            "cv_max_size_mb": cv_max // (1024 * 1024),
        }
        return render(request, self.template_name, context)

    def post(self, request, pk):
        """Guardar evaluación manual (Cumple/No cumple por criterio) y recalcular score."""
        candidate, ats_form = self._get_candidate_and_form(request, pk)
        if not candidate or not ats_form:
            return redirect("ats_dashboard")
        criteria = list(ats_form.criteria.all())
        prefix = "criterion_"
        for c in criteria:
            key = f"{prefix}{c.id}"
            raw = request.POST.get(key)
            cumple = str(raw).strip().lower() in ("1", "true", "si", "cumple", "on")
            resp, _ = ATSCandidateCriterionResponse.objects.get_or_create(
                candidate=candidate,
                criterion=c,
                defaults={"cumple": cumple},
            )
            if resp.cumple != cumple:
                resp.cumple = cumple
                resp.save(update_fields=["cumple"])
        # Score = (criterios que cumple / total criterios) * 100
        total = len(criteria)
        if total > 0:
            cumplidos = ATSCandidateCriterionResponse.objects.filter(
                candidate=candidate,
                criterion__in=criteria,
                cumple=True,
            ).count()
            new_score = round((cumplidos / total) * 100)
            candidate.score = new_score
            if new_score >= 70:
                candidate.status = Candidate.STATUS_APTO
            elif new_score < 40:
                candidate.status = Candidate.STATUS_NO_APTO
            else:
                candidate.status = Candidate.STATUS_REVISION
            candidate.save(update_fields=["score", "status"])
        messages.success(request, "Evaluación guardada. Score actualizado.")
        return redirect("ats_candidate_detail", pk=pk)


class ATSCandidateUploadCVView(LoginRequiredMixin, View):
    """Subir o reemplazar el CV de un candidato (PDF/DOCX)."""
    login_url = reverse_lazy("ats_plataforma")
    http_method_names = ["post"]

    def post(self, request, pk):
        ats_client = getattr(request.user, "ats_client", None)
        if not ats_client:
            return redirect("ats_dashboard")
        candidate = get_object_or_404(Candidate, pk=pk, client=ats_client)
        cv_file = request.FILES.get("cv_file")
        if not cv_file:
            messages.error(request, "Selecciona un archivo (PDF o DOCX).")
            return redirect("ats_candidate_detail", pk=pk)
        # Validar extensión y tamaño (mismo que formulario público)
        from django.conf import settings as django_settings
        allowed = getattr(django_settings, "ATS_FORM_PUBLIC_ALLOWED_EXTENSIONS", ["pdf", "doc", "docx"])
        max_size = getattr(django_settings, "ATS_FORM_PUBLIC_MAX_FILE_SIZE", 10 * 1024 * 1024)
        ext = (cv_file.name or "").split(".")[-1].lower()
        if ext not in allowed:
            messages.error(request, f"Formato no permitido. Usa: {', '.join(allowed)}.")
            return redirect("ats_candidate_detail", pk=pk)
        if cv_file.size > max_size:
            messages.error(request, f"El archivo es demasiado grande. Máximo {max_size // (1024*1024)} MB.")
            return redirect("ats_candidate_detail", pk=pk)
        from django.core.files.base import ContentFile
        name = cv_file.name or "cv.pdf"
        if candidate.cv_file:
            candidate.cv_file.delete(save=False)
        candidate.cv_file.save(name, ContentFile(cv_file.read()), save=True)
        messages.success(request, "CV cargado correctamente. Ya puedes analizarlo con IA si lo deseas.")
        return redirect("ats_candidate_detail", pk=pk)


class ATSCandidateAnalyzeCVView(LoginRequiredMixin, View):
    """Ejecutar análisis del CV con IA y guardar score/habilidades en la BD."""
    login_url = reverse_lazy("ats_plataforma")
    http_method_names = ["post"]

    def post(self, request, pk):
        ats_client = getattr(request.user, "ats_client", None)
        if not ats_client:
            return redirect("ats_dashboard")
        candidate = get_object_or_404(Candidate, pk=pk, client=ats_client)
        subscription = _get_or_create_subscription(request.user)
        if not subscription_can(subscription, "cvs_scan"):
            messages.error(request, "No tienes análisis de CV disponibles o tu plan no incluye escaneo con IA.")
            return redirect("ats_candidate_detail", pk=pk)
        from mi_app.services.cv_analysis import run_cv_analysis_and_save
        result = run_cv_analysis_and_save(candidate)
        if not result.get("ok"):
            messages.error(request, result.get("error", "Error al analizar el CV."))
            return redirect("ats_candidate_detail", pk=pk)
        subscription.increment_cvs_used()
        messages.success(request, "Análisis completado. Score y habilidades guardados según el perfil de la vacante.")
        return redirect("ats_candidate_detail", pk=pk)


class ATSCandidateSendEmailView(LoginRequiredMixin, View):
    """Envía correo al candidato: apto para reclutar o no seleccionado (rechazo)."""
    login_url = reverse_lazy("ats_plataforma")
    http_method_names = ["post"]

    def post(self, request, pk):
        ats_client = getattr(request.user, "ats_client", None)
        if not ats_client:
            return redirect("ats_dashboard")
        candidate = get_object_or_404(Candidate, pk=pk, client=ats_client)
        if not (candidate.email or "").strip():
            messages.error(request, "Este candidato no tiene correo registrado.")
            return redirect("ats_candidate_detail", pk=pk)
        email_type = (request.POST.get("email_type") or "").strip().lower()
        if email_type not in ("apto", "rechazo"):
            messages.error(request, "Tipo de correo no válido.")
            return redirect("ats_candidate_detail", pk=pk)
        subscription = _get_or_create_subscription(request.user)
        custom_message = (request.POST.get("custom_message") or "").strip() or None
        if custom_message and not subscription_can(subscription, "custom_email_message"):
            custom_message = None
        ok = send_email_to_candidate(ats_client, candidate, email_type, custom_message)
        if ok:
            messages.success(
                request,
                "Correo enviado al candidato." if email_type == "apto" else "Correo de no seleccionado enviado al candidato.",
            )
        else:
            messages.error(request, "No se pudo enviar el correo. Revisa la configuración de correo en Config. correo.")
        return redirect("ats_candidate_detail", pk=pk)


class ATSCandidateExportView(LoginRequiredMixin, View):
    """Exporta candidatos a CSV o Excel. Solo planes PRO y ENTERPRISE."""
    login_url = reverse_lazy("ats_plataforma")
    http_method_names = ["get"]

    def get(self, request):
        ats_client = getattr(request.user, "ats_client", None)
        if not ats_client:
            return redirect("ats_dashboard")
        subscription = _get_or_create_subscription(request.user)
        if not subscription_can(subscription, "export_candidates"):
            messages.error(request, "La exportación de candidatos está disponible en planes Pro y Enterprise.")
            return redirect(reverse("ats_dashboard") + "?section=candidatos")
        fmt = (request.GET.get("format") or "csv").strip().lower()
        if fmt not in ("csv", "xlsx"):
            fmt = "csv"
        base_qs = Candidate.objects.filter(client=ats_client).select_related("vacancy")
        q = (request.GET.get("q") or "").strip()
        status_filter = request.GET.get("status", "")
        vacancy_id = request.GET.get("vacancy", "")
        qs = base_qs
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(email__icontains=q))
        if status_filter and status_filter in (Candidate.STATUS_APTO, Candidate.STATUS_REVISION, Candidate.STATUS_NO_APTO):
            qs = qs.filter(status=status_filter)
        if vacancy_id:
            try:
                qs = qs.filter(vacancy_id=int(vacancy_id))
            except ValueError:
                pass
        qs = qs.order_by("-analysis_date", "-id")[:5000]
        rows = []
        for c in qs:
            rows.append({
                "nombre": c.name or "",
                "email": c.email or "",
                "vacante": c.vacancy.title if c.vacancy else "",
                "score": c.score,
                "estado": c.status,
                "match_percentage": c.match_percentage if c.match_percentage is not None else "",
                "fecha_analisis": c.analysis_date.strftime("%Y-%m-%d %H:%M") if c.analysis_date else "",
            })
        if fmt == "csv":
            return self._response_csv(rows)
        try:
            import openpyxl
        except ImportError:
            messages.error(request, "Exportación Excel no disponible. Instala openpyxl o usa formato CSV.")
            return redirect(reverse("ats_dashboard") + "?section=candidatos")
        return self._response_xlsx(rows)

    def _response_csv(self, rows):
        import csv
        import io
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="candidatos.csv"'
        response.write("\ufeff")
        writer = csv.writer(response)
        headers = ["nombre", "email", "vacante", "score", "estado", "match_percentage", "fecha_analisis"]
        writer.writerow(headers)
        for r in rows:
            writer.writerow([r.get(h, "") for h in headers])
        return response

    def _response_xlsx(self, rows):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Candidatos"
        headers = ["nombre", "email", "vacante", "score", "estado", "match_percentage", "fecha_analisis"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        for row_idx, r in enumerate(rows, 2):
            for col_idx, h in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=r.get(h, ""))
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="candidatos.xlsx"'
        wb.save(response)
        return response


def _get_client_or_403(request):
    """Devuelve ATSClient del usuario o None si no tiene; para vistas que requieren cliente."""
    ats_client = getattr(request.user, "ats_client", None)
    if not ats_client:
        return None
    return ats_client


# --- Formularios ATS (crear, editar, listar, ver envíos) ---

class ATSFormListView(LoginRequiredMixin, ListView):
    """Lista de formularios del cliente con enlace público y envíos."""
    model = ATSForm
    template_name = "ats/form_list.html"
    login_url = reverse_lazy("ats_plataforma")
    context_object_name = "forms"

    def get_queryset(self):
        client = _get_client_or_403(self.request)
        if not client:
            return ATSForm.objects.none()
        return (
            ATSForm.objects.filter(client=client)
            .select_related("vacancy")
            .prefetch_related("fields")
            .annotate(submissions_count=Count("submissions"))
            .order_by("-updated_at")
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["ats_client"] = _get_client_or_403(self.request)
        context["subscription"] = _get_or_create_subscription(self.request.user)
        context["ats_page"] = "formularios"
        return context


class ATSFormCreateView(LoginRequiredMixin, View):
    """Crear formulario (nombre, descripción, vacante) y redirigir a editar para añadir campos."""
    login_url = reverse_lazy("ats_plataforma")

    def get(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("ats_dashboard")
        form = ATSFormCreateEditForm(initial={"vacancy": None})
        form.fields["vacancy"].queryset = client.vacancies.all()
        return render(request, "ats/form_create.html", {
            "form": form,
            "ats_client": client,
            "subscription": _get_or_create_subscription(request.user),
            "ats_page": "formularios",
        })

    def post(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("ats_dashboard")
        form = ATSFormCreateEditForm(request.POST)
        form.fields["vacancy"].queryset = client.vacancies.all()
        if form.is_valid():
            ats_form = form.save(commit=False)
            ats_form.client = client
            ats_form.save()
            messages.success(request, "Formulario creado. Añade los campos y guarda.")
            return redirect("ats_form_edit", pk=ats_form.pk)
        return render(request, "ats/form_create.html", {
            "form": form,
            "ats_client": client,
            "subscription": _get_or_create_subscription(request.user),
            "ats_page": "formularios",
        })


class ATSFormEditView(LoginRequiredMixin, View):
    """Editar formulario, sus campos y criterios de evaluación manual (formsets)."""
    login_url = reverse_lazy("ats_plataforma")

    def get(self, request, pk):
        client = _get_client_or_403(request)
        if not client:
            return redirect("ats_dashboard")
        ats_form = get_object_or_404(ATSForm, pk=pk, client=client)
        form = ATSFormCreateEditForm(instance=ats_form)
        form.fields["vacancy"].queryset = client.vacancies.all()
        formset = get_ats_form_field_formset(extra=1, form_instance=ats_form)
        criteria_formset = get_ats_form_criterion_formset(extra=1, form_instance=ats_form)
        return render(request, "ats/form_edit.html", {
            "ats_form": ats_form,
            "form": form,
            "formset": formset,
            "criteria_formset": criteria_formset,
            "ats_client": client,
            "subscription": _get_or_create_subscription(request.user),
            "ats_page": "formularios",
        })

    def post(self, request, pk):
        client = _get_client_or_403(request)
        if not client:
            return redirect("ats_dashboard")
        ats_form = get_object_or_404(ATSForm, pk=pk, client=client)
        form = ATSFormCreateEditForm(request.POST, instance=ats_form)
        form.fields["vacancy"].queryset = client.vacancies.all()
        formset = get_ats_form_field_formset(extra=1, form_instance=ats_form, data=request.POST, files=request.FILES)
        criteria_formset = get_ats_form_criterion_formset(extra=1, form_instance=ats_form, data=request.POST)
        if form.is_valid() and formset.is_valid() and criteria_formset.is_valid():
            form.save()
            formset.save(commit=False)
            for obj in formset.new_objects:
                obj.form = ats_form
                obj.save()
            for obj, _ in formset.changed_objects:
                obj.save()
            for obj in formset.deleted_objects:
                obj.delete()
            criteria_formset.save(commit=False)
            for obj in criteria_formset.new_objects:
                obj.form = ats_form
                obj.save()
            for obj, _ in criteria_formset.changed_objects:
                obj.save()
            for obj in criteria_formset.deleted_objects:
                obj.delete()
            messages.success(request, "Formulario guardado.")
            return redirect("ats_form_list")
        return render(request, "ats/form_edit.html", {
            "ats_form": ats_form,
            "form": form,
            "formset": formset,
            "criteria_formset": criteria_formset,
            "ats_client": client,
            "subscription": _get_or_create_subscription(request.user),
            "ats_page": "formularios",
        })


class ATSFormDeleteView(LoginRequiredMixin, View):
    """Eliminar formulario."""
    login_url = reverse_lazy("ats_plataforma")

    def post(self, request, pk):
        client = _get_client_or_403(request)
        if not client:
            return redirect("ats_dashboard")
        ats_form = get_object_or_404(ATSForm, pk=pk, client=client)
        ats_form.delete()
        messages.success(request, "Formulario eliminado.")
        return redirect("ats_form_list")


class ATSFormSubmissionsView(LoginRequiredMixin, ListView):
    """Envíos recibidos de un formulario."""
    model = ATSFormSubmission
    template_name = "ats/form_submissions.html"
    login_url = reverse_lazy("ats_plataforma")
    context_object_name = "submissions"

    def get_queryset(self):
        client = _get_client_or_403(self.request)
        if not client:
            return ATSFormSubmission.objects.none()
        ats_form = get_object_or_404(ATSForm, pk=self.kwargs["pk"], client=client)
        self.ats_form = ats_form
        return ATSFormSubmission.objects.filter(form=ats_form).prefetch_related("files").order_by("-submitted_at")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["ats_form"] = getattr(self, "ats_form", None)
        context["ats_client"] = _get_client_or_403(self.request)
        context["subscription"] = _get_or_create_subscription(self.request.user)
        context["ats_page"] = "formularios"
        return context


class ATSFormPublicView(View):
    """Formulario público: GET muestra el formulario, POST recibe el envío."""
    template_name = "ats/form_public.html"
    thank_you_template = "ats/form_public_thankyou.html"

    def get(self, request, uuid):
        ats_form = get_object_or_404(ATSForm, uuid=uuid, is_active=True)
        has_email_field = ats_form.fields.filter(field_type=ATSFormField.FIELD_EMAIL).exists()
        return render(request, self.template_name, {
            "ats_form": ats_form,
            "ats_form_has_email_field": has_email_field,
        })

    def post(self, request, uuid):
        ats_form = get_object_or_404(ATSForm, uuid=uuid, is_active=True)
        # Rate limit por IP + formulario
        ip = request.META.get("REMOTE_ADDR", "") or "unknown"
        cache_key = f"ats_form_submit:{ip}:{uuid}"
        count = cache.get(cache_key, 0)
        max_count = getattr(settings, "ATS_FORM_PUBLIC_RATE_LIMIT_COUNT", 5)
        if count >= max_count:
            return render(request, self.template_name, {
                "ats_form": ats_form,
                "ats_form_has_email_field": ats_form.fields.filter(field_type=ATSFormField.FIELD_EMAIL).exists(),
                "form_error": "Has alcanzado el límite de envíos. Intenta de nuevo más tarde.",
            })
        # Validación de archivos (tamaño y extensión)
        max_size = getattr(settings, "ATS_FORM_PUBLIC_MAX_FILE_SIZE", 10 * 1024 * 1024)
        allowed_ext = getattr(settings, "ATS_FORM_PUBLIC_ALLOWED_EXTENSIONS", ["pdf", "doc", "docx"])
        def _check_file(f):
            if f.size > max_size:
                return False, f"El archivo {getattr(f, 'name', '')} supera el tamaño máximo permitido ({max_size // (1024*1024)} MB)."
            ext = (f.name or "").rsplit(".", 1)[-1].lower() if "." in (f.name or "") else ""
            if ext and allowed_ext and ext not in allowed_ext:
                return False, f"Solo se permiten archivos: {', '.join(allowed_ext)}."
            return True, None
        payload = {}
        files_to_save = []
        submitter_email = ""
        for field in ats_form.fields.all().order_by("order", "id"):
            key = f"field_{field.id}"
            if field.field_type == ATSFormField.FIELD_FILE:
                f = request.FILES.get(key)
                if f and field.required or f:
                    ok, err = _check_file(f)
                    if not ok:
                        return render(request, self.template_name, {
                            "ats_form": ats_form,
                            "ats_form_has_email_field": ats_form.fields.filter(field_type=ATSFormField.FIELD_EMAIL).exists(),
                            "form_error": err,
                        })
                    files_to_save.append((field, f))
                    payload[field.label] = f.name
            else:
                val = request.POST.get(key, "").strip()
                if val or field.required:
                    payload[field.label] = val
                    if field.field_type == ATSFormField.FIELD_EMAIL and not submitter_email:
                        submitter_email = val
        if getattr(ats_form, "request_cv", False):
            cv_file = request.FILES.get("cv_file")
            if cv_file:
                ok, err = _check_file(cv_file)
                if not ok:
                    return render(request, self.template_name, {
                        "ats_form": ats_form,
                        "ats_form_has_email_field": ats_form.fields.filter(field_type=ATSFormField.FIELD_EMAIL).exists(),
                        "form_error": err,
                    })
                files_to_save.append((None, cv_file))
                payload["CV"] = cv_file.name
        submitter_email = submitter_email or request.POST.get("submitter_email", "").strip()
        if submitter_email and "Correo electrónico" not in payload and "Email" not in payload:
            payload["Correo electrónico"] = submitter_email
        submission = ATSFormSubmission.objects.create(
            form=ats_form,
            payload=payload,
            submitter_email=submitter_email or "",
        )
        for field, uploaded_file in files_to_save:
            ATSFormSubmissionFile.objects.create(
                submission=submission,
                form_field=field,
                file=uploaded_file,
                original_name=uploaded_file.name,
            )
        # Si el formulario está ligado a una vacante, crear candidato y una sola notificación "Nuevo candidato"
        if ats_form.vacancy_id:
            _create_candidate_from_submission(submission, payload, submitter_email)
            if submission.candidate_id:
                notify_ats_client(
                    ats_form.client,
                    ATSNotification.TYPE_CANDIDATE,
                    "Nuevo candidato",
                    message=f"{submission.candidate.name} — Formulario «{ats_form.name}».",
                    link=reverse("ats_candidate_detail", args=[submission.candidate.pk]),
                    request=request,
                )
        else:
            # Sin vacante: notificación solo de envío
            notify_ats_client(
                ats_form.client,
                ATSNotification.TYPE_SUBMISSION,
                "Nuevo envío de formulario",
                message=f"Formulario «{ats_form.name}»: {submitter_email or 'Sin correo'}.",
                link=reverse("ats_form_submissions", args=[ats_form.pk]),
                request=request,
            )
        # Rate limit: incrementar tras envío correcto
        timeout = getattr(settings, "ATS_FORM_PUBLIC_RATE_LIMIT_SECONDS", 3600)
        cache.set(cache_key, count + 1, timeout=timeout)
        return redirect("ats_form_public_thanks", uuid=uuid)


class ATSFormPublicThanksView(View):
    """Página de agradecimiento tras enviar el formulario público (GET; se llega por redirect tras POST)."""
    template_name = "ats/form_public_thankyou.html"

    def get(self, request, uuid):
        ats_form = get_object_or_404(ATSForm, uuid=uuid, is_active=True)
        return render(request, self.template_name, {"ats_form": ats_form})


def _create_candidate_from_submission(submission, payload, submitter_email):
    """Crea un Candidato a partir de un envío de formulario y lo vincula. Usado en POST público y en backfill. Respeta límite de candidatos del plan."""
    ats_form = submission.form
    if not ats_form.vacancy_id:
        return None
    client = ats_form.client
    subscription = getattr(client.user, "ats_subscription", None)
    current_count = Candidate.objects.filter(client=client).count()
    if not subscription_can_add_candidate(subscription, current_count):
        return None
    submitter_email = submitter_email or submission.submitter_email or ""
    name = (
        payload.get("Nombre")
        or payload.get("Nombre completo")
        or payload.get("Nombre y apellidos")
        or payload.get("name")
        or (submitter_email.split("@")[0] if submitter_email else "Postulante")
    )
    if not name or len(str(name)) > 255:
        name = (str(name) if name else "Postulante")[:255]
    candidate = Candidate.objects.create(
        client=ats_form.client,
        vacancy=ats_form.vacancy,
        name=name,
        email=submitter_email or "",
        status=Candidate.STATUS_REVISION,
        score=0,
        explanation_text="Postulación recibida por formulario. Pendiente de análisis de CV con IA.",
    )
    submission.candidate = candidate
    submission.save(update_fields=["candidate"])
    # Si el envío incluyó un archivo (CV), copiarlo al candidato para poder procesarlo con IA después
    # Preferir el archivo del campo "Solicitar CV" (form_field=None); si no, el primer archivo adjunto
    cv_attachment = submission.files.filter(form_field__isnull=True).first() or submission.files.first()
    if cv_attachment and cv_attachment.file:
        from django.core.files.base import ContentFile
        try:
            name = cv_attachment.original_name or "cv_adjunto"
            if not name or "." not in name:
                name = name or "cv_adjunto"
            candidate.cv_file.save(name, ContentFile(cv_attachment.file.read()), save=True)
        except Exception:
            pass
    return candidate


class ATSEmailConfigView(LoginRequiredMixin, FormView):
    """Configuración de correo: notificaciones y correo de la empresa (conectar inbox)."""
    template_name = "ats/email_config.html"
    form_class = ATSEmailConfigForm
    login_url = reverse_lazy("ats_plataforma")
    success_url = reverse_lazy("ats_email_config")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        client = _get_client_or_403(self.request)
        if not client:
            return kwargs
        config, _ = ATSClientEmailConfig.objects.get_or_create(client=client)
        kwargs["instance"] = config
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["ats_client"] = _get_client_or_403(self.request)
        context["subscription"] = _get_or_create_subscription(self.request.user)
        context["ats_page"] = "correo"
        return context

    def form_valid(self, form):
        client = _get_client_or_403(self.request)
        if not client:
            return redirect("ats_dashboard")
        form.instance.client = client
        form.save()
        messages.success(self.request, "Configuración de correo guardada.")
        return super().form_valid(form)


class ATSProfileConfigView(LoginRequiredMixin, FormView):
    """Configurar cuenta: foto de perfil, nombre, teléfono, empresa."""
    template_name = "ats/profile_config.html"
    form_class = ATSProfileForm
    login_url = reverse_lazy("ats_plataforma")
    success_url = reverse_lazy("ats_profile_config")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        client = _get_client_or_403(self.request)
        if client:
            kwargs["instance"] = client
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["ats_client"] = _get_client_or_403(self.request)
        context["subscription"] = _get_or_create_subscription(self.request.user)
        context["ats_page"] = "perfil"
        return context

    def form_valid(self, form):
        client = _get_client_or_403(self.request)
        if not client:
            return redirect("ats_dashboard")
        form.save()
        messages.success(self.request, "Cuenta actualizada.")
        return redirect(self.get_success_url())


class ATSVacancyCreateView(LoginRequiredMixin, View):
    """Crear vacante (puesto) desde Reclutamiento."""
    login_url = reverse_lazy("ats_plataforma")

    def get(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("ats_dashboard")
        subscription = _get_or_create_subscription(request.user)
        vacancy_count = Vacancy.objects.filter(client=client).count()
        if not subscription_can_add_vacancy(subscription, vacancy_count):
            messages.error(request, "Has alcanzado el límite de vacantes de tu plan (Gratuito: 2 vacantes). Mejora tu plan para crear más.")
            return redirect(reverse("ats_dashboard") + "?section=reclutamiento")
        return render(request, "ats/vacancy_form.html", {
            "form": ATSVacancyForm(),
            "ats_client": client,
            "subscription": subscription,
            "ats_page": "reclutamiento",
        })

    def post(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("ats_dashboard")
        subscription = _get_or_create_subscription(request.user)
        vacancy_count = Vacancy.objects.filter(client=client).count()
        if not subscription_can_add_vacancy(subscription, vacancy_count):
            messages.error(request, "Has alcanzado el límite de vacantes de tu plan. Mejora tu plan para crear más.")
            return redirect(reverse("ats_dashboard") + "?section=reclutamiento")
        form = ATSVacancyForm(request.POST)
        if form.is_valid():
            vacancy = form.save(commit=False)
            vacancy.client = client
            vacancy.save()
            messages.success(request, "Vacante creada. Ya puedes asociar candidatos y formularios a este puesto.")
            return redirect(reverse("ats_dashboard") + "?section=reclutamiento")
        return render(request, "ats/vacancy_form.html", {
            "form": form,
            "ats_client": client,
            "subscription": subscription,
            "ats_page": "reclutamiento",
        })


class ATSVacancyEditView(LoginRequiredMixin, View):
    """Editar vacante (título, descripción, perfil para análisis con IA)."""
    login_url = reverse_lazy("ats_plataforma")

    def get(self, request, pk):
        client = _get_client_or_403(request)
        if not client:
            return redirect("ats_dashboard")
        vacancy = get_object_or_404(Vacancy, pk=pk, client=client)
        form = ATSVacancyForm(instance=vacancy)
        return render(request, "ats/vacancy_form.html", {
            "form": form,
            "ats_client": client,
            "subscription": _get_or_create_subscription(request.user),
            "ats_page": "reclutamiento",
            "vacancy": vacancy,
            "is_edit": True,
        })

    def post(self, request, pk):
        client = _get_client_or_403(request)
        if not client:
            return redirect("ats_dashboard")
        vacancy = get_object_or_404(Vacancy, pk=pk, client=client)
        form = ATSVacancyForm(request.POST, instance=vacancy)
        if form.is_valid():
            form.save()
            messages.success(request, "Vacante actualizada.")
            return redirect(reverse("ats_dashboard") + "?section=reclutamiento")
        return render(request, "ats/vacancy_form.html", {
            "form": form,
            "ats_client": client,
            "subscription": _get_or_create_subscription(request.user),
            "ats_page": "reclutamiento",
            "vacancy": vacancy,
            "is_edit": True,
        })


class ATSVacancyDeleteView(LoginRequiredMixin, View):
    """Eliminar vacante."""
    login_url = reverse_lazy("ats_plataforma")

    def post(self, request, pk):
        client = _get_client_or_403(request)
        if not client:
            return redirect("ats_dashboard")
        vacancy = get_object_or_404(Vacancy, pk=pk, client=client)
        vacancy.delete()
        messages.success(request, "Vacante eliminada.")
        return redirect(reverse("ats_dashboard") + "?section=reclutamiento")


class ATSCVAnalysisConfigView(LoginRequiredMixin, View):
    """Configuración por defecto del análisis de CV con IA (perfil e instrucciones)."""
    login_url = reverse_lazy("ats_plataforma")

    def get(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("ats_dashboard")
        config, _ = CVAnalysisConfig.objects.get_or_create(client=client)
        form = CVAnalysisConfigForm(instance=config)
        return render(request, "ats/cv_analysis_config.html", {
            "form": form,
            "ats_client": client,
            "subscription": _get_or_create_subscription(request.user),
            "ats_page": "analisis_cv",
        })

    def post(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("ats_dashboard")
        config, _ = CVAnalysisConfig.objects.get_or_create(client=client)
        form = CVAnalysisConfigForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            messages.success(request, "Configuración de análisis de CV guardada.")
            return redirect("ats_cv_analysis_config")
        return render(request, "ats/cv_analysis_config.html", {
            "form": form,
            "ats_client": client,
            "subscription": _get_or_create_subscription(request.user),
            "ats_page": "analisis_cv",
        })


class ATSBackfillCandidatesFromSubmissionsView(LoginRequiredMixin, View):
    """Crea Candidatos desde envíos de formularios que tienen vacante pero aún no tienen candidato vinculado."""
    login_url = reverse_lazy("ats_plataforma")
    http_method_names = ["post"]

    def post(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("ats_dashboard")
        pending = ATSFormSubmission.objects.filter(
            form__client=client,
            form__vacancy__isnull=False,
            candidate__isnull=True,
        ).select_related("form", "form__vacancy")
        created = 0
        for sub in pending:
            if _create_candidate_from_submission(sub, sub.payload, sub.submitter_email):
                created += 1
        if created:
            messages.success(
                request,
                f"Se crearon {created} candidato(s) desde envíos de formularios. Ya aparecen en Reclutamiento y Candidatos.",
            )
            notify_ats_client(
                client,
                ATSNotification.TYPE_CANDIDATE,
                "Candidatos creados desde envíos",
                message=f"Se crearon {created} candidato(s) desde formularios. Revisa la sección Candidatos.",
                link=reverse("ats_dashboard") + "?section=candidatos",
                request=request,
            )
        else:
            messages.info(request, "No había envíos pendientes de vincular. Todos los envíos con vacante ya tienen candidato.")
        return redirect(reverse("ats_dashboard") + "?section=reclutamiento")


class ATSNotificationGoView(LoginRequiredMixin, View):
    """Marca una notificación como leída y redirige a su enlace."""
    login_url = reverse_lazy("ats_plataforma")
    http_method_names = ["get"]

    def get(self, request, pk):
        client = _get_client_or_403(request)
        if not client:
            return redirect("ats_dashboard")
        notification = get_object_or_404(ATSNotification, pk=pk, client=client)
        notification.read = True
        notification.save(update_fields=["read"])
        link = (notification.link or "").strip()
        if link and not link.startswith("http"):
            link = request.build_absolute_uri(link)
        return redirect(link or reverse("ats_dashboard"))


class ATSNotificationMarkAllReadView(LoginRequiredMixin, View):
    """Marca todas las notificaciones del cliente como leídas."""
    login_url = reverse_lazy("ats_plataforma")
    http_method_names = ["post"]

    def post(self, request):
        client = _get_client_or_403(request)
        if client:
            ATSNotification.objects.filter(client=client, read=False).update(read=True)
        ref = request.META.get("HTTP_REFERER") or reverse("ats_dashboard")
        return redirect(ref)


class ATSNotificationPanelView(LoginRequiredMixin, View):
    """Panel de notificaciones: lista paginada con opción de marcar todas como leídas."""
    login_url = reverse_lazy("ats_plataforma")
    template_name = "ats/notification_panel.html"
    paginate_by = 25

    def get(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("ats_dashboard")
        qs = ATSNotification.objects.filter(client=client).order_by("-created_at")
        paginator = Paginator(qs, self.paginate_by)
        page_number = request.GET.get("page", 1)
        page = paginator.get_page(page_number)
        unread_count = ATSNotification.objects.filter(client=client, read=False).count()
        return render(request, self.template_name, {
            "page_obj": page,
            "notifications": page.object_list,
            "ats_client": client,
            "ats_unread_count": unread_count,
            "ats_page": "notificaciones",
            "subscription": _get_or_create_subscription(request.user),
        })

    def post(self, request):
        """Marcar todas como leídas y redirigir al panel."""
        client = _get_client_or_403(request)
        if client:
            ATSNotification.objects.filter(client=client, read=False).update(read=True)
        return redirect("ats_notification_panel")


# --- Panel de administración ATS (solo staff) ---

class ATSAdminDashboardView(StaffRequiredMixin, View):
    """Panel para soporte: listar clientes ATS y sus suscripciones; cambiar plan."""
    template_name = "ats/admin/dashboard.html"
    paginate_by = 25

    def get(self, request):
        qs = ATSClient.objects.select_related("user").order_by("-created_at")
        search_q = (request.GET.get("q") or "").strip()
        if search_q:
            qs = qs.filter(
                Q(company_name__icontains=search_q)
                | Q(contact_name__icontains=search_q)
                | Q(user__email__icontains=search_q)
            )
        paginator = Paginator(qs, self.paginate_by)
        page_number = request.GET.get("page", 1)
        page_obj = paginator.get_page(page_number)
        client_ids = [c.user_id for c in page_obj.object_list]
        subs_by_user = {
            sub.user_id: sub
            for sub in Subscription.objects.filter(user_id__in=client_ids).select_related("user")
        }
        rows = []
        for client in page_obj.object_list:
            rows.append({"client": client, "subscription": subs_by_user.get(client.user_id)})
        plan_choices = [
            (Subscription.PLAN_FREE, "Gratuito"),
            (Subscription.PLAN_PRO, "Pro"),
            (Subscription.PLAN_ENTERPRISE, "Enterprise"),
        ]
        pending_requests = list(
            PlanChangeRequest.objects.filter(status=PlanChangeRequest.STATUS_PENDING)
            .select_related("client")
            .order_by("-created_at")[:20]
        )
        # Uso de IA (tokens): totales para mostrar en admin y enlace a LangSmith
        now = timezone.now()
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        llm_usage_this_month = LLMUsageLog.objects.filter(created_at__gte=month_start).aggregate(
            total=Sum("total_tokens"), runs=Count("id")
        )
        llm_usage_all_time = LLMUsageLog.objects.aggregate(total=Sum("total_tokens"), runs=Count("id"))
        import os
        langsmith_configured = bool(os.environ.get("LANGSMITH_API_KEY", "").strip())
        langsmith_project = os.environ.get("LANGSMITH_PROJECT", "default")
        langsmith_runs = []
        if langsmith_configured:
            try:
                from langsmith import Client
                client = Client()
                for run in client.list_runs(project_name=langsmith_project, limit=15):
                    extra = getattr(run, "extra", None) or {}
                    usage = extra.get("usage") or {}
                    langsmith_runs.append({
                        "name": getattr(run, "name", "—"),
                        "start_time": getattr(run, "start_time", None),
                        "total_tokens": usage.get("total_tokens"),
                        "inputs": getattr(run, "inputs", None),
                    })
            except Exception:
                langsmith_runs = []
        return render(request, self.template_name, {
            "rows": rows,
            "page_obj": page_obj,
            "search_q": search_q,
            "plan_choices": plan_choices,
            "pending_requests": pending_requests,
            "llm_usage_this_month": llm_usage_this_month.get("total") or 0,
            "llm_usage_this_month_runs": llm_usage_this_month.get("runs") or 0,
            "llm_usage_all_time": llm_usage_all_time.get("total") or 0,
            "llm_usage_all_time_runs": llm_usage_all_time.get("runs") or 0,
            "langsmith_configured": langsmith_configured,
            "langsmith_project": langsmith_project,
            "langsmith_runs": langsmith_runs,
            "ats_page": "administracion",
            "ats_client": getattr(request.user, "ats_client", None),
            "ats_header_title": "Administración ATS",
        })


class ATSAdminChangePlanView(StaffRequiredMixin, View):
    """POST: aplicar un plan a una suscripción (tras validar pago)."""
    def post(self, request):
        subscription_id = request.POST.get("subscription_id")
        plan_id = (request.POST.get("plan") or "").strip().upper()
        if not subscription_id or plan_id not in (Subscription.PLAN_FREE, Subscription.PLAN_PRO, Subscription.PLAN_ENTERPRISE):
            messages.error(request, "Datos inválidos.")
            return redirect("ats_admin_dashboard")
        subscription = get_object_or_404(Subscription, pk=subscription_id)
        old_plan = subscription.plan
        ok = apply_plan_to_subscription(subscription, plan_id)
        if not ok:
            messages.error(request, "No se pudo aplicar el plan.")
            return redirect("ats_admin_dashboard")
        messages.success(request, f"Plan actualizado a {subscription.get_plan_display()} para {subscription.user.email}.")
        client = getattr(subscription.user, "ats_client", None)
        if client:
            PlanChangeRequest.objects.filter(
                client=client, to_plan=plan_id, status=PlanChangeRequest.STATUS_PENDING
            ).update(status=PlanChangeRequest.STATUS_DONE)
            notify_ats_client(
                client,
                ATSNotification.TYPE_PLAN,
                "Plan actualizado",
                message=f"Tu plan ha sido actualizado a {subscription.get_plan_display()}.",
                link=reverse("ats_dashboard") + "?section=cuenta",
                request=request,
            )
        return redirect("ats_admin_dashboard")


class ATSAdminSetLangSmithView(StaffRequiredMixin, View):
    """POST: asignar proyecto LangSmith a un cliente (desde el panel ATS administración)."""
    http_method_names = ["post"]

    def post(self, request):
        client_id = request.POST.get("client_id")
        langsmith_project = (request.POST.get("langsmith_project") or "").strip()[:100]
        if not client_id:
            messages.error(request, "Falta el cliente.")
            return redirect("ats_admin_dashboard")
        client = get_object_or_404(ATSClient, pk=client_id)
        client.langsmith_project = langsmith_project
        client.save(update_fields=["langsmith_project"])
        if langsmith_project:
            messages.success(request, f"Proyecto LangSmith «{langsmith_project}» asignado a {client.company_name}.")
        else:
            messages.success(request, f"Proyecto LangSmith quitado para {client.company_name}. Se usará el global.")
        return redirect("ats_admin_dashboard")


class ATSStaffAccountView(StaffRequiredMixin, View):
    """Mi cuenta para el administrador (staff): datos de usuario y enlace a cambiar contraseña."""
    template_name = "ats/admin/mi_cuenta.html"

    def get(self, request):
        if request.user.is_staff and not getattr(request.user, "ats_client", None):
            return render(request, self.template_name, {
                "ats_page": "mi_cuenta_staff",
                "ats_client": None,
                "ats_header_title": "Mi cuenta",
            })
        return redirect("ats_admin_dashboard")


class ATSPasswordChangeView(StaffRequiredMixin, AuthPasswordChangeView):
    """Cambio de contraseña para el administrador ATS (mismo layout)."""
    template_name = "ats/admin/password_change_form.html"
    success_url = reverse_lazy("ats_staff_account")
    login_url = reverse_lazy("ats_plataforma")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["ats_page"] = "mi_cuenta_staff"
        context["ats_client"] = None
        context["ats_header_title"] = "Cambiar contraseña"
        return context


# --- Recuperar contraseña (cliente y staff) ---

class ATSPasswordResetView(AuthPasswordResetView):
    template_name = "ats/auth/password_reset_form.html"
    success_url = reverse_lazy("ats_password_reset_done")
    email_template_name = "ats/auth/email/password_reset_email.html"
    subject_template_name = "ats/auth/email/password_reset_subject.txt"


class ATSPasswordResetDoneView(AuthPasswordResetDoneView):
    template_name = "ats/auth/password_reset_done.html"


class ATSPasswordResetConfirmView(AuthPasswordResetConfirmView):
    template_name = "ats/auth/password_reset_confirm.html"
    success_url = reverse_lazy("ats_password_reset_complete")
    post_reset_login = False


class ATSPasswordResetCompleteView(AuthPasswordResetCompleteView):
    template_name = "ats/auth/password_reset_complete.html"

