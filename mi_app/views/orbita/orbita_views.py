"""
Vistas para el producto ATS (Applicant Tracking System) de Star Path.
- Página de oferta / producto: "Ven y prueba nuestro ATS"
- Plataforma ATS: login, registro y dashboard para clientes.
- Dashboard: candidatos, habilidades (en detalle), cuenta/billing.
"""
import json
import logging
import uuid as uuid_lib

from django.core.paginator import Paginator
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import TemplateView, View
from django.contrib.auth import login, logout, get_user_model
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import (
    PasswordChangeView as AuthPasswordChangeView,
    PasswordResetView as AuthPasswordResetView,
    PasswordResetDoneView as AuthPasswordResetDoneView,
    PasswordResetConfirmView as AuthPasswordResetConfirmView,
    PasswordResetCompleteView as AuthPasswordResetCompleteView,
)
from django.urls import reverse_lazy, reverse

from django.conf import settings
from django.core.cache import cache
from django.core.mail import send_mail
from django.core.mail.backends.smtp import EmailBackend
from django.contrib import messages
from django.http import HttpResponseForbidden, HttpResponseRedirect, HttpResponse, JsonResponse, FileResponse
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, FormView
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.decorators.cache import never_cache

from mi_app.views.orbita.forms import (
    ATSRegisterForm,
    ATSLoginForm,
    ATSFormCreateEditForm,
    get_orbita_form_field_formset,
    get_orbita_form_criterion_formset,
    ATSEmailConfigForm,
    ATSProfileForm,
    ATSAvatarUploadForm,
    ATSVacancyForm,
    WorkforceAreaForm,
    WorkforcePositionForm,
    WorkforcePlanForm,
    CVAnalysisConfigForm,
)
from django.db.models import Count, Q, Sum  # Count for annotate, Q for filter
from datetime import timedelta

from mi_app.models import (
    ATSClient,
    Subscription,
    Candidate,
    Vacancy,
    CVAnalysisConfig,
    ATSForm,
    ATSFormField,
    ATSFormCriterion,
    ATSFormSubmission,
    ATSFormSubmissionFile,
    ATSCandidateCriterionResponse,
    ATSClientEmailConfig,
    ATSNotification,
    PlanChangeRequest,
    LLMUsageLog,
    WorkforceArea,
    WorkforcePosition,
    WorkforcePlan,
)
from mi_app.orbita_plans import (
    get_all_plans,
    get_plan_capabilities_display,
    get_plan_candidates_limit,
    get_plan_vacancies_limit,
    subscription_can,
    subscription_can_add_candidate,
    subscription_can_add_vacancy,
    apply_plan_to_subscription,
    get_subscription_module_flags,
    subscription_module_enabled,
)
from mi_app.orbita_notifications import notify_orbita_client, notify_support_plan_change, notify_support_account_deletion_request, send_email_to_candidate
from mi_app.services.form_submissions import (
    create_submission_once,
    normalize_submitter_email,
)

User = get_user_model()
logger = logging.getLogger(__name__)


def _is_valid_uuid(value):
    try:
        uuid_lib.UUID(str(value))
    except (TypeError, ValueError, AttributeError):
        return False
    return True


class StaffRequiredMixin(LoginRequiredMixin):
    """Solo usuarios con is_staff pueden acceder. Para el panel de administración ATS."""
    login_url = reverse_lazy("orbita_plataforma")

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        if not request.user.is_staff:
            return HttpResponseForbidden("No tienes permiso para acceder a esta página.")
        return super().dispatch(request, *args, **kwargs)


class OrbitaModuleRequiredMixin:
    """Bloquea vistas de cliente cuando soporte apagó el módulo en la suscripción."""
    module_required = None

    def dispatch(self, request, *args, **kwargs):
        if self.module_required and not _request_module_enabled(request, self.module_required):
            messages.warning(request, "Este módulo no está habilitado para tu cuenta.")
            fallback = _first_enabled_dashboard_url(request)
            if fallback:
                return redirect(fallback)
            return HttpResponseForbidden("No tienes módulos habilitados para tu cuenta.")
        return super().dispatch(request, *args, **kwargs)


class ATSProductoView(TemplateView):
    """Página que ofrece el producto ATS: Low Code, identificación de talento, formularios, OpenAI, match."""
    template_name = "orbita/producto_orbita.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["plans"] = get_all_plans()
        return context


class ATSPlataformaView(View):
    """Página de acceso: login y registro. Si ya está autenticado, redirige al dashboard."""
    template_name = "orbita/plataforma.html"

    def get(self, request):
        if request.user.is_authenticated:
            return redirect("orbita_dashboard")
        active_tab = request.GET.get("tab", "")
        register_data = request.session.pop("orbita_register_invalid_data", None)
        login_data = request.session.pop("orbita_login_invalid_data", None)
        login_error = request.session.pop("orbita_login_error", "")
        register_form = ATSRegisterForm(register_data) if register_data else ATSRegisterForm()
        login_form = ATSLoginForm(request, data=login_data) if login_data else ATSLoginForm(request, initial={
            "username": request.session.pop("orbita_login_username", "")
        })
        if register_data:
            register_form.is_valid()
            active_tab = "register"
        if login_data:
            login_form.is_valid()
            active_tab = "login"
        if login_error:
            active_tab = "login"
        return render(request, self.template_name, {
            "login_form": login_form,
            "register_form": register_form,
            "active_tab": active_tab,
            "login_error": login_error,
        })

    def post(self, request):
        # El POST se maneja en las URLs específicas (login o register)
        return self.get(request)


class ATSRegisterView(View):
    """Registro de nuevo cliente ATS."""
    template_name = "orbita/plataforma.html"
    success_url = reverse_lazy("orbita_dashboard")

    def get(self, request):
        if request.user.is_authenticated:
            return redirect("orbita_dashboard")
        return redirect("orbita_plataforma")

    def post(self, request):
        if request.user.is_authenticated:
            return redirect("orbita_dashboard")
        form = ATSRegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user, backend="django.contrib.auth.backends.ModelBackend")
            return redirect(self.success_url)
        request.session["orbita_register_invalid_data"] = {
            key: value
            for key, value in request.POST.items()
            if key not in ("csrfmiddlewaretoken", "password1", "password2")
        }
        return redirect(reverse("orbita_plataforma") + "?tab=register")


class ATSLoginView(View):
    """Inicio de sesión ATS (correo + contraseña)."""
    template_name = "orbita/plataforma.html"
    success_url = reverse_lazy("orbita_dashboard")

    def get(self, request):
        if request.user.is_authenticated:
            if request.user.is_staff:
                return redirect("orbita_admin_dashboard")
            return redirect("orbita_dashboard")
        return redirect("orbita_plataforma")

    def post(self, request):
        if request.user.is_authenticated:
            if request.user.is_staff:
                return redirect("orbita_admin_dashboard")
            return redirect("orbita_dashboard")
        form = ATSLoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user, backend="django.contrib.auth.backends.ModelBackend")
            if request.GET.get("next"):
                return redirect(request.GET["next"])
            if user.is_staff:
                return redirect("orbita_admin_dashboard")
            return redirect(self.success_url)
        if request.POST.get("username") and request.POST.get("password"):
            request.session["orbita_login_username"] = request.POST.get("username", "")
            request.session["orbita_login_error"] = "Correo o contraseña incorrectos."
        else:
            request.session["orbita_login_invalid_data"] = {
                key: value
                for key, value in request.POST.items()
                if key not in ("csrfmiddlewaretoken", "password")
            }
        return redirect(reverse("orbita_plataforma") + "?tab=login")


class ATSLogoutView(View):
    """Cerrar sesión y volver a la página de plataforma."""
    def get(self, request):
        logout(request)
        return redirect("orbita_plataforma")

    def post(self, request):
        logout(request)
        return redirect("orbita_plataforma")


def _get_or_create_subscription(user):
    """Crea suscripción FREE si el usuario no tiene (ej. clientes creados antes de este módulo)."""
    sub, _ = Subscription.objects.get_or_create(
        user=user,
        defaults={
            "plan": Subscription.PLAN_FREE,
            "cvs_limit": 3,
            "active": True,
        },
    )
    return sub


def _request_module_enabled(request, module):
    if request.user.is_staff and not getattr(request.user, "ats_client", None):
        return True
    if not request.user.is_authenticated:
        return False
    subscription = _get_or_create_subscription(request.user)
    return subscription_module_enabled(subscription, module)


def _first_enabled_dashboard_url(request):
    modules_by_section = (
        ("workforce", "workforce"),
        ("candidates", "candidatos"),
        ("vacancies", "reclutamiento"),
        ("account", "cuenta"),
    )
    for module, section in modules_by_section:
        if _request_module_enabled(request, module):
            if module == "workforce":
                return reverse("orbita_workforce_dashboard")
            if module == "account":
                return reverse("orbita_profile_config")
            return reverse("orbita_dashboard") + f"?section={section}"
    return None


def user_can_process_cv(user):
    """
    Verifica si el usuario puede procesar un CV más (límite de plan).
    Usar antes de aceptar subida de CV / llamada a LLM; si falla, devolver 403 o mensaje claro.
    """
    sub = _get_or_create_subscription(user)
    return sub.can_process_cv


class ATSDashboardView(LoginRequiredMixin, TemplateView):
    """Panel del cliente ATS: KPIs, candidatos (filtro), gráfica, reclutamiento (vacantes), cuenta."""
    template_name = "orbita/dashboard.html"
    login_url = reverse_lazy("orbita_plataforma")
    allowed_sections = {"candidatos", "reclutamiento", "cuenta"}

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return super().dispatch(request, *args, **kwargs)
        section = request.GET.get("section", "candidatos")
        if section == "cuenta":
            return redirect("orbita_profile_config")
        if section not in self.allowed_sections:
            return redirect(reverse("orbita_dashboard") + "?section=candidatos")
        vacancy_filter = request.GET.get("vacancy", "")
        if vacancy_filter and not _is_valid_uuid(vacancy_filter):
            clean_query = request.GET.copy()
            clean_query.pop("vacancy", None)
            if not clean_query.get("section"):
                clean_query["section"] = "candidatos"
            query_string = clean_query.urlencode()
            clean_url = reverse("orbita_dashboard")
            return redirect(f"{clean_url}?{query_string}" if query_string else clean_url)
        required = {
            "candidatos": "candidates",
            "reclutamiento": "vacancies",
            "cuenta": "account",
        }[section]
        if not _request_module_enabled(request, required):
            messages.warning(request, "Este módulo no está habilitado para tu cuenta.")
            fallback = _first_enabled_dashboard_url(request)
            if fallback:
                return redirect(fallback)
            return HttpResponseForbidden("No tienes módulos habilitados para tu cuenta.")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        orbita_client = getattr(self.request.user, "ats_client", None)
        context["orbita_client"] = orbita_client
        subscription = _get_or_create_subscription(self.request.user)
        context["subscription"] = subscription
        section = self.request.GET.get("section", "candidatos")
        context["dashboard_section"] = section
        context["orbita_page"] = section
        # Planes con sus capacidades (vistas/procedimientos/actividades por plan)
        all_plans = get_all_plans()
        for p in all_plans:
            p["capabilities_display"] = get_plan_capabilities_display(p["id"])
        context["available_plans"] = all_plans
        context["subscription_plan_capabilities"] = get_plan_capabilities_display(subscription.plan)
        context["subscription_can_process_cv"] = subscription_can(subscription, "cvs_scan")
        context["subscription_can_export_candidates"] = subscription_can(subscription, "export_candidates")

        if orbita_client:
            base_qs = Candidate.objects.filter(client=orbita_client).select_related("vacancy")
            # KPIs globales (sin filtros) para secciones generales
            context["kpi_total_global"] = base_qs.count()
            context["kpi_aptos_global"] = base_qs.filter(status=Candidate.STATUS_APTO).count()
            context["kpi_revision_global"] = base_qs.filter(status=Candidate.STATUS_REVISION).count()
            context["kpi_no_aptos_global"] = base_qs.filter(status=Candidate.STATUS_NO_APTO).count()
            context["kpi_cvs_used"] = subscription.cvs_used
            context["kpi_cvs_limit"] = subscription.cvs_limit
            context["kpi_candidates_limit"] = get_plan_candidates_limit(subscription.plan)
            context["kpi_vacancies_limit"] = get_plan_vacancies_limit(subscription.plan)
            context["kpi_vacancy_count"] = Vacancy.objects.filter(client=orbita_client).count()
            # Filtros
            q = (self.request.GET.get("q") or "").strip()
            status_filter = self.request.GET.get("status", "")
            vacancy_public_id = self.request.GET.get("vacancy", "")
            qs = base_qs
            if q:
                qs = qs.filter(Q(name__icontains=q) | Q(email__icontains=q))
            if status_filter and status_filter in (Candidate.STATUS_APTO, Candidate.STATUS_REVISION, Candidate.STATUS_NO_APTO):
                qs = qs.filter(status=status_filter)
            selected_vacancy = None
            if vacancy_public_id:
                selected_vacancy = Vacancy.objects.filter(client=orbita_client, public_id=vacancy_public_id).first()
                if selected_vacancy:
                    qs = qs.filter(vacancy=selected_vacancy)
            # KPIs de vista candidatos (sí respetan filtros activos, incluida vacante)
            context["kpi_total"] = qs.count()
            context["kpi_aptos"] = qs.filter(status=Candidate.STATUS_APTO).count()
            context["kpi_revision"] = qs.filter(status=Candidate.STATUS_REVISION).count()
            context["kpi_no_aptos"] = qs.filter(status=Candidate.STATUS_NO_APTO).count()
            # Datos para gráfica (por estado, filtrados)
            context["chart_aptos"] = context["kpi_aptos"]
            context["chart_revision"] = context["kpi_revision"]
            context["chart_no_aptos"] = context["kpi_no_aptos"]
            context["candidates"] = qs.order_by("-score", "-analysis_date").prefetch_related("skill_evaluations")[:200]
            context["filter_q"] = q
            context["filter_status"] = status_filter
            context["filter_vacancy"] = str(selected_vacancy.public_id) if selected_vacancy else ""
            context["selected_vacancy"] = selected_vacancy
            # Vacantes (para filtro y sección Reclutamiento)
            context["vacancies"] = Vacancy.objects.filter(client=orbita_client).annotate(
                candidates_count=Count("candidates")
            ).order_by("-created_at")
            # Envíos de formularios con vacante pero sin candidato (para botón "Crear candidatos desde envíos")
            context["pending_submissions_count"] = ATSFormSubmission.objects.filter(
                form__client=orbita_client,
                form__vacancy__isnull=False,
                candidate__isnull=True,
            ).count()
        else:
            context["kpi_total"] = context["kpi_aptos"] = context["kpi_revision"] = context["kpi_no_aptos"] = 0
            context["kpi_total_global"] = context["kpi_aptos_global"] = context["kpi_revision_global"] = context["kpi_no_aptos_global"] = 0
            context["kpi_cvs_used"] = subscription.cvs_used
            context["kpi_cvs_limit"] = subscription.cvs_limit
            context["kpi_candidates_limit"] = get_plan_candidates_limit(subscription.plan)
            context["kpi_vacancies_limit"] = get_plan_vacancies_limit(subscription.plan)
            context["kpi_vacancy_count"] = 0
            context["chart_aptos"] = context["chart_revision"] = context["chart_no_aptos"] = 0
            context["candidates"] = []
            context["filter_q"] = context["filter_status"] = context["filter_vacancy"] = ""
            context["selected_vacancy"] = None
            context["vacancies"] = []
            context["pending_submissions_count"] = 0
            context["subscription_can_export_candidates"] = subscription_can(subscription, "export_candidates")
        return context


class ATSChangePlanView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """
    El cliente solicita cambio de plan. No se aplica el cambio aquí: se envía correo
    a soporte para validar, gestionar pago externo y activar el plan desde el admin.
    """
    login_url = reverse_lazy("orbita_plataforma")
    http_method_names = ["post"]
    module_required = "account"

    def post(self, request):
        plan_id = (request.POST.get("plan") or "").strip().upper()
        if plan_id not in (Subscription.PLAN_FREE, Subscription.PLAN_PRO, Subscription.PLAN_ENTERPRISE):
            messages.warning(request, "Plan no válido.")
            return redirect(reverse("orbita_dashboard") + "?section=cuenta")
        subscription = _get_or_create_subscription(request.user)
        old_plan_id = subscription.plan
        if old_plan_id == plan_id:
            messages.info(request, "Ya tienes ese plan seleccionado.")
            return redirect(reverse("orbita_dashboard") + "?section=cuenta")
        orbita_client = _get_client_or_403(request)
        # Enviar correo a soporte para validar, cobro externo y activar desde admin
        notify_support_plan_change(request.user, orbita_client, old_plan_id, plan_id)
        if orbita_client:
            PlanChangeRequest.objects.create(
                client=orbita_client,
                from_plan=old_plan_id,
                to_plan=plan_id,
                status=PlanChangeRequest.STATUS_PENDING,
            )
        plan_name = dict(Subscription.PLAN_CHOICES).get(plan_id, plan_id)
        if orbita_client:
            notify_orbita_client(
                orbita_client,
                ATSNotification.TYPE_PLAN,
                "Solicitud de cambio de plan",
                message=f"Solicitud enviada para pasar a {plan_name}. Soporte validará y te contactará; al gestionar el pago se activará tu nuevo plan.",
                link=reverse("orbita_dashboard") + "?section=cuenta",
                request=request,
            )
        messages.success(
            request,
            "Solicitud enviada. Soporte validará el cambio y te contactará; una vez gestionado el pago, se activará tu nuevo plan desde administración.",
        )
        return redirect(reverse("orbita_dashboard") + "?section=cuenta")


class ATSRequestAccountDeletionView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """Solicitar baja de cuenta: envía correo a soporte y redirige a Configurar cuenta."""
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "account"

    def get(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        return render(request, "orbita/request_account_deletion.html", {
            "orbita_client": client,
            "orbita_page": "perfil",
        })

    def post(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        notify_support_account_deletion_request(client)
        messages.success(
            request,
            "Tu solicitud de baja ha sido enviada a soporte. Te contactaremos para confirmar.",
        )
        return redirect("orbita_profile_config")


class ATSCandidateDetailView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """Detalle de candidato: score, estado, explicación, habilidades y evaluación manual (Cumple/No cumple)."""
    template_name = "orbita/candidate_detail.html"
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "candidates"

    def _get_candidate_and_form(self, request, public_id):
        orbita_client = getattr(request.user, "ats_client", None)
        if not orbita_client:
            return None, None, False
        try:
            candidate = Candidate.objects.prefetch_related(
                "skill_evaluations",
                "criterion_responses",
                "form_submissions__files",
                "form_submissions__form",
            ).get(public_id=public_id, client=orbita_client)
        except Candidate.DoesNotExist:
            return None, None, True
        orbita_form = None
        first_sub = candidate.form_submissions.first()
        if first_sub:
            orbita_form = first_sub.form
        return candidate, orbita_form, False

    def get(self, request, public_id):
        candidate, orbita_form, was_deleted = self._get_candidate_and_form(request, public_id)
        if was_deleted:
            return redirect("orbita_dashboard")
        if not candidate:
            return render(request, self.template_name, {"candidate": None, "orbita_page": "candidatos"})
        criteria = list(orbita_form.criteria.all()) if orbita_form else []
        response_map = {r.criterion_id: r.cumple for r in candidate.criterion_responses.filter(criterion__in=criteria)}
        evaluation_criteria_with_response = [(c, response_map.get(c.id, False)) for c in criteria]
        form_submission = candidate.form_submissions.first()
        subscription = _get_or_create_subscription(request.user)
        from django.conf import settings as django_settings
        cv_max = getattr(django_settings, "ORBITA_FORM_PUBLIC_MAX_FILE_SIZE", 10 * 1024 * 1024)

        chat_session = None
        chat_conversation = []
        if form_submission:
            try:
                chat_session = form_submission.chat_session
            except Exception:
                chat_session = None
            if chat_session and orbita_form:
                from mi_app.views.orbita.form_chat_views import _build_steps
                steps = _build_steps(orbita_form)
                answers = chat_session.answers or {}
                for i, s in enumerate(steps):
                    val = answers.get(s["id"])
                    chat_conversation.append({
                        "label": s["label"],
                        "type": s["type"],
                        "answer": val,
                        "answered": val is not None,
                        "is_current": i == chat_session.current_step and chat_session.status != "completed",
                    })

        context = {
            "candidate": candidate,
            "orbita_client": getattr(request.user, "ats_client", None),
            "subscription": subscription,
            "orbita_page": "candidatos",
            "form_submission": form_submission,
            "evaluation_criteria": criteria,
            "evaluation_criteria_with_response": evaluation_criteria_with_response,
            "subscription_can_process_cv": subscription_can(subscription, "cvs_scan"),
            "subscription_can_custom_email": subscription_can(subscription, "custom_email_message"),
            "kpi_cvs_used": subscription.cvs_used if subscription else 0,
            "kpi_cvs_limit": subscription.cvs_limit if subscription else 0,
            "cv_max_size_mb": cv_max // (1024 * 1024),
            "chat_session": chat_session,
            "chat_conversation": chat_conversation,
        }
        return render(request, self.template_name, context)

    def post(self, request, public_id):
        """Guardar evaluación manual (Cumple/No cumple por criterio) y recalcular score."""
        candidate, orbita_form, was_deleted = self._get_candidate_and_form(request, public_id)
        if was_deleted:
            return redirect("orbita_dashboard")
        if not candidate or not orbita_form:
            return redirect("orbita_dashboard")
        criteria = list(orbita_form.criteria.all())
        prefix = "criterion_"
        for c in criteria:
            key = f"{prefix}{c.id}"
            raw = request.POST.get(key)
            cumple = str(raw).strip().lower() in ("1", "true", "si", "cumple", "on")
            resp, _ = ATSCandidateCriterionResponse.objects.get_or_create(
                candidate=candidate,
                criterion=c,
                defaults={"cumple": cumple},
            )
            if resp.cumple != cumple:
                resp.cumple = cumple
                resp.save(update_fields=["cumple"])
        # Score ponderado: suma de valores cumplidos / suma total de valores * 100
        total_valor = sum(max(0, min(100, int(getattr(c, "score_value", 0) or 0))) for c in criteria)
        if total_valor > 0:
            cumplidos_ids = set(
                ATSCandidateCriterionResponse.objects.filter(
                    candidate=candidate,
                    criterion__in=criteria,
                    cumple=True,
                ).values_list("criterion_id", flat=True)
            )
            puntos = sum(
                max(0, min(100, int(getattr(c, "score_value", 0) or 0)))
                for c in criteria
                if c.id in cumplidos_ids
            )
            new_score = round((puntos / total_valor) * 100)
            candidate.score = new_score
            if new_score >= 70:
                candidate.status = Candidate.STATUS_APTO
            elif new_score < 40:
                candidate.status = Candidate.STATUS_NO_APTO
            else:
                candidate.status = Candidate.STATUS_REVISION
            candidate.save(update_fields=["score", "status"])
        else:
            candidate.score = 0
            candidate.status = Candidate.STATUS_REVISION
            candidate.save(update_fields=["score", "status"])
        messages.success(request, "Evaluación guardada. Score actualizado.")
        return redirect("orbita_candidate_detail", public_id=candidate.public_id)


class ATSCandidateUploadCVView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """Subir o reemplazar el CV de un candidato (PDF/DOCX)."""
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "candidates"
    http_method_names = ["post"]

    def post(self, request, public_id):
        orbita_client = getattr(request.user, "ats_client", None)
        if not orbita_client:
            return redirect("orbita_dashboard")
        candidate = get_object_or_404(Candidate, public_id=public_id, client=orbita_client)
        cv_file = request.FILES.get("cv_file")
        if not cv_file:
            messages.error(request, "Selecciona un archivo (PDF o DOCX).")
            return redirect("orbita_candidate_detail", public_id=candidate.public_id)
        # Validar extensión y tamaño (mismo que formulario público)
        from django.conf import settings as django_settings
        allowed = getattr(django_settings, "ORBITA_FORM_PUBLIC_ALLOWED_EXTENSIONS", ["pdf", "doc", "docx"])
        max_size = getattr(django_settings, "ORBITA_FORM_PUBLIC_MAX_FILE_SIZE", 10 * 1024 * 1024)
        ext = (cv_file.name or "").split(".")[-1].lower()
        if ext not in allowed:
            messages.error(request, f"Formato no permitido. Usa: {', '.join(allowed)}.")
            return redirect("orbita_candidate_detail", public_id=candidate.public_id)
        if cv_file.size > max_size:
            messages.error(request, f"El archivo es demasiado grande. Máximo {max_size // (1024*1024)} MB.")
            return redirect("orbita_candidate_detail", public_id=candidate.public_id)
        from django.core.files.base import ContentFile
        name = cv_file.name or "cv.pdf"
        if candidate.cv_file:
            candidate.cv_file.delete(save=False)
        candidate.cv_file.save(name, ContentFile(cv_file.read()), save=True)
        messages.success(request, "CV cargado correctamente. Ya puedes analizarlo con IA si lo deseas.")
        return redirect("orbita_candidate_detail", public_id=candidate.public_id)


class ATSCandidateAnalyzeCVView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """Ejecutar análisis del CV con IA y guardar score/habilidades en la BD."""
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "cv_analysis"
    http_method_names = ["post"]

    def post(self, request, public_id):
        orbita_client = getattr(request.user, "ats_client", None)
        if not orbita_client:
            return redirect("orbita_dashboard")
        candidate = get_object_or_404(Candidate, public_id=public_id, client=orbita_client)
        cv_config, _ = CVAnalysisConfig.objects.get_or_create(client=orbita_client)
        if not cv_config.enabled:
            messages.error(
                request,
                "El análisis de CV con IA está deshabilitado en tu configuración. Actívalo en Config. análisis CV.",
            )
            return redirect("orbita_candidate_detail", public_id=candidate.public_id)
        if candidate.vacancy_id and not getattr(candidate.vacancy, "ai_enabled", False):
            messages.error(
                request,
                "La IA está desactivada para la vacante de este candidato. Actívala en Vacantes > Editar vacante.",
            )
            return redirect("orbita_candidate_detail", public_id=candidate.public_id)
        subscription = _get_or_create_subscription(request.user)
        if not subscription_can(subscription, "cvs_scan"):
            messages.error(request, "No tienes análisis de CV disponibles o tu plan no incluye escaneo con IA.")
            return redirect("orbita_candidate_detail", public_id=candidate.public_id)
        from mi_app.services.cv_analysis import run_cv_analysis_and_save
        result = run_cv_analysis_and_save(candidate)
        if not result.get("ok"):
            messages.error(request, result.get("error", "Error al analizar el CV."))
            return redirect("orbita_candidate_detail", public_id=candidate.public_id)
        subscription.increment_cvs_used()
        messages.success(request, "Análisis completado. Score y habilidades guardados según el perfil de la vacante.")
        return redirect("orbita_candidate_detail", public_id=candidate.public_id)


class ATSCandidateSendEmailView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """Envía correo al candidato: apto para reclutar o no seleccionado (rechazo)."""
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "candidates"
    http_method_names = ["post"]

    def post(self, request, public_id):
        orbita_client = getattr(request.user, "ats_client", None)
        if not orbita_client:
            return redirect("orbita_dashboard")
        candidate = get_object_or_404(Candidate, public_id=public_id, client=orbita_client)
        if not (candidate.email or "").strip():
            messages.error(request, "Este candidato no tiene correo registrado.")
            return redirect("orbita_candidate_detail", public_id=candidate.public_id)
        email_type = (request.POST.get("email_type") or "").strip().lower()
        if email_type not in ("apto", "rechazo"):
            messages.error(request, "Tipo de correo no válido.")
            return redirect("orbita_candidate_detail", public_id=candidate.public_id)
        subscription = _get_or_create_subscription(request.user)
        custom_message = (request.POST.get("custom_message") or "").strip() or None
        if custom_message and not subscription_can(subscription, "custom_email_message"):
            custom_message = None
        ok = send_email_to_candidate(orbita_client, candidate, email_type, custom_message)
        if ok:
            messages.success(
                request,
                "Correo enviado al candidato." if email_type == "apto" else "Correo de no seleccionado enviado al candidato.",
            )
        else:
            messages.error(request, "No se pudo enviar el correo. Revisa la configuración de correo en Config. correo.")
        return redirect("orbita_candidate_detail", public_id=candidate.public_id)


class ATSCandidateExportView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """Exporta candidatos a CSV/Excel o sus CVs en ZIP. Solo planes PRO y ENTERPRISE."""
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "candidates"
    http_method_names = ["get"]

    def get(self, request):
        orbita_client = getattr(request.user, "ats_client", None)
        if not orbita_client:
            return redirect("orbita_dashboard")
        subscription = _get_or_create_subscription(request.user)
        if not subscription_can(subscription, "export_candidates"):
            messages.error(request, "La exportación de candidatos está disponible en planes Pro y Enterprise.")
            return redirect(reverse("orbita_dashboard") + "?section=candidatos")
        fmt = (request.GET.get("format") or "csv").strip().lower()
        if fmt not in ("csv", "xlsx", "zip"):
            fmt = "csv"
        base_qs = Candidate.objects.filter(client=orbita_client).select_related("vacancy")
        q = (request.GET.get("q") or "").strip()
        status_filter = request.GET.get("status", "")
        vacancy_public_id = request.GET.get("vacancy", "")
        qs = base_qs
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(email__icontains=q))
        if status_filter and status_filter in (Candidate.STATUS_APTO, Candidate.STATUS_REVISION, Candidate.STATUS_NO_APTO):
            qs = qs.filter(status=status_filter)
        if vacancy_public_id and _is_valid_uuid(vacancy_public_id):
            vacancy = Vacancy.objects.filter(client=orbita_client, public_id=vacancy_public_id).first()
            if vacancy:
                qs = qs.filter(vacancy=vacancy)
        qs = qs.order_by("-analysis_date", "-id")[:5000]
        if fmt == "zip":
            return self._response_zip(qs, orbita_client)
        rows = []
        for c in qs:
            rows.append({
                "nombre": c.name or "",
                "email": c.email or "",
                "vacante": c.vacancy.title if c.vacancy else "",
                "score": c.score,
                "estado": c.status,
                "match_percentage": c.match_percentage if c.match_percentage is not None else "",
                "fecha_analisis": c.analysis_date.strftime("%Y-%m-%d %H:%M") if c.analysis_date else "",
            })
        if fmt == "csv":
            return self._response_csv(rows)
        try:
            import openpyxl
        except ImportError:
            messages.error(request, "Exportación Excel no disponible. Instala openpyxl o usa formato CSV.")
            return redirect(reverse("orbita_dashboard") + "?section=candidatos")
        return self._response_xlsx(rows)

    def _response_zip(self, candidates, orbita_client):
        import os
        import tempfile
        import zipfile
        from django.utils.text import slugify

        tmp = tempfile.TemporaryFile()
        added = 0
        with zipfile.ZipFile(tmp, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            for candidate in candidates:
                if not candidate.cv_file:
                    continue
                try:
                    original = os.path.basename(candidate.cv_file.name) or "cv.pdf"
                    stem, ext = os.path.splitext(original)
                    ext = ext or ".pdf"
                    safe_name = slugify(candidate.name or f"candidato-{candidate.pk}") or f"candidato-{candidate.pk}"
                    archive_name = f"cvs/{candidate.pk}-{safe_name}{ext.lower()}"
                    with candidate.cv_file.open("rb") as fh:
                        zf.writestr(archive_name, fh.read())
                    added += 1
                except Exception as exc:
                    logger.warning("No se pudo agregar CV al ZIP candidate=%s: %s", candidate.pk, exc)

            if added == 0:
                zf.writestr("sin_cvs.txt", "No hay CVs disponibles para los filtros seleccionados.")

        tmp.seek(0)
        company = slugify(getattr(orbita_client, "company_name", "") or "orbita") or "orbita"
        response = FileResponse(tmp, as_attachment=True, filename=f"cvs-{company}.zip")
        response["Content-Type"] = "application/zip"
        return response

    def _response_csv(self, rows):
        import csv
        import io
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="candidatos.csv"'
        response.write("\ufeff")
        writer = csv.writer(response)
        headers = ["nombre", "email", "vacante", "score", "estado", "match_percentage", "fecha_analisis"]
        writer.writerow(headers)
        for r in rows:
            writer.writerow([r.get(h, "") for h in headers])
        return response

    def _response_xlsx(self, rows):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Candidatos"
        headers = ["nombre", "email", "vacante", "score", "estado", "match_percentage", "fecha_analisis"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        for row_idx, r in enumerate(rows, 2):
            for col_idx, h in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=r.get(h, ""))
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="candidatos.xlsx"'
        wb.save(response)
        return response


class ATSCandidateDeleteView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """POST: elimina un candidato."""
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "candidates"
    http_method_names = ["post"]

    def post(self, request, public_id):
        orbita_client = getattr(request.user, "ats_client", None)
        if not orbita_client:
            return redirect("orbita_dashboard")
        candidate = get_object_or_404(Candidate, public_id=public_id, client=orbita_client)
        candidate.delete()
        messages.success(request, "Candidato eliminado correctamente.")
        return redirect(reverse("orbita_dashboard") + "?section=candidatos")


def _get_client_or_403(request):
    """Devuelve ATSClient del usuario o None si no tiene; para vistas que requieren cliente."""
    orbita_client = getattr(request.user, "ats_client", None)
    if not orbita_client:
        return None
    return orbita_client


def _sync_criteria_from_form_fields(orbita_form):
    """
    Garantiza que los campos del formulario (excepto archivo) aparezcan como criterios manuales.
    Crea criterios vinculados si faltan y mantiene etiqueta/orden sincronizados.
    """
    form_fields = list(
        orbita_form.fields.exclude(field_type=ATSFormField.FIELD_FILE).order_by("order", "id")
    )
    field_ids = {f.id for f in form_fields}
    ATSFormCriterion.objects.filter(form=orbita_form, source_form_field__isnull=True).delete()
    ATSFormCriterion.objects.filter(form=orbita_form).exclude(source_form_field_id__in=field_ids).delete()

    existing_by_source = {
        c.source_form_field_id: c
        for c in ATSFormCriterion.objects.filter(form=orbita_form, source_form_field__isnull=False)
    }
    for idx, f in enumerate(form_fields):
        c = existing_by_source.get(f.id)
        if c:
            updates = []
            if c.label != f.label:
                c.label = f.label
                updates.append("label")
            if c.order != idx:
                c.order = idx
                updates.append("order")
            if updates:
                c.save(update_fields=updates)
            continue
        ATSFormCriterion.objects.create(
            form=orbita_form,
            source_form_field=f,
            label=f.label,
            score_value=100,
            order=idx,
        )


# --- Formularios ATS (crear, editar, listar, ver envíos) ---

class ATSFormListView(OrbitaModuleRequiredMixin, LoginRequiredMixin, ListView):
    """Lista de formularios del cliente con enlace público y envíos."""
    model = ATSForm
    template_name = "orbita/form_list.html"
    login_url = reverse_lazy("orbita_plataforma")
    context_object_name = "forms"
    module_required = "forms"

    def get_queryset(self):
        client = _get_client_or_403(self.request)
        if not client:
            return ATSForm.objects.none()
        return (
            ATSForm.objects.filter(client=client)
            .select_related("vacancy")
            .prefetch_related("fields")
            .annotate(submissions_count=Count("submissions"))
            .order_by("-updated_at")
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["orbita_client"] = _get_client_or_403(self.request)
        context["subscription"] = _get_or_create_subscription(self.request.user)
        context["orbita_page"] = "formularios"
        forms_qs = context.get("forms") or self.get_queryset()
        context["total_submissions"] = sum(getattr(f, "submissions_count", 0) for f in forms_qs)
        context["total_active"] = sum(1 for f in forms_qs if f.is_active)
        return context


class ATSFormCreateView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """Crear formulario (nombre, descripción, vacante) y redirigir a editar para añadir campos."""
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "forms"

    def get(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        form = ATSFormCreateEditForm(initial={"vacancy": None})
        form.fields["vacancy"].queryset = client.vacancies.all()
        return render(request, "orbita/form_create.html", {
            "form": form,
            "orbita_client": client,
            "subscription": _get_or_create_subscription(request.user),
            "orbita_page": "formularios",
        })

    def post(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        form = ATSFormCreateEditForm(request.POST)
        form.fields["vacancy"].queryset = client.vacancies.all()
        if form.is_valid():
            orbita_form = form.save(commit=False)
            orbita_form.client = client
            orbita_form.save()
            messages.success(request, "Formulario creado. Añade los campos y guarda.")
            return redirect("orbita_form_edit", pk=orbita_form.pk)
        return render(request, "orbita/form_create.html", {
            "form": form,
            "orbita_client": client,
            "subscription": _get_or_create_subscription(request.user),
            "orbita_page": "formularios",
        })


class ATSFormEditView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """Editar formulario, sus campos y criterios de evaluación manual (formsets)."""
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "forms"

    def get(self, request, pk):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        orbita_form = get_object_or_404(ATSForm, pk=pk, client=client)
        _sync_criteria_from_form_fields(orbita_form)
        form = ATSFormCreateEditForm(instance=orbita_form)
        form.fields["vacancy"].queryset = client.vacancies.all()
        formset = get_orbita_form_field_formset(extra=0, form_instance=orbita_form)
        criteria_formset = get_orbita_form_criterion_formset(extra=0, form_instance=orbita_form)
        return render(request, "orbita/form_edit.html", {
            "orbita_form": orbita_form,
            "form": form,
            "formset": formset,
            "criteria_formset": criteria_formset,
            "orbita_client": client,
            "subscription": _get_or_create_subscription(request.user),
            "orbita_page": "formularios",
        })

    def post(self, request, pk):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        orbita_form = get_object_or_404(ATSForm, pk=pk, client=client)
        if request.POST.get("autosave_criteria") == "1":
            return self._autosave_criteria(request, orbita_form)
        form = ATSFormCreateEditForm(request.POST, instance=orbita_form)
        form.fields["vacancy"].queryset = client.vacancies.all()
        formset = get_orbita_form_field_formset(extra=0, form_instance=orbita_form, data=request.POST, files=request.FILES)
        criteria_formset = get_orbita_form_criterion_formset(extra=0, form_instance=orbita_form, data=request.POST)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save(commit=False)
            for obj in formset.new_objects:
                obj.form = orbita_form
                obj.save()
            for obj, _ in formset.changed_objects:
                obj.save()
            for obj in formset.deleted_objects:
                obj.delete()

            # Mapa índice-formset -> pk de campo guardado, para aplicar score automático de criterios nuevos.
            field_index_to_pk = {}
            for idx, field_form in enumerate(formset.forms):
                cleaned = getattr(field_form, "cleaned_data", None) or {}
                if not cleaned or cleaned.get("DELETE"):
                    continue
                instance = field_form.instance
                if not instance.pk:
                    continue
                if instance.field_type == ATSFormField.FIELD_FILE:
                    continue
                field_index_to_pk[str(idx)] = instance.pk

            _sync_criteria_from_form_fields(orbita_form)
            self._apply_posted_existing_criteria_scores(request, orbita_form)

            # Campos nuevos mostrados por AJAX: auto_score_field_<índice_formset>
            # Se aplica después del sync, porque el criterio aún no existe antes de sincronizar.
            for key, value in request.POST.items():
                if not key.startswith("auto_score_field_"):
                    continue
                field_idx = key[len("auto_score_field_"):]
                field_pk = field_index_to_pk.get(field_idx)
                if not field_pk:
                    continue
                try:
                    score_value = int(value)
                except (TypeError, ValueError):
                    score_value = 100
                score_value = max(0, min(100, score_value))
                ATSFormCriterion.objects.filter(
                    form=orbita_form,
                    source_form_field_id=field_pk,
                ).update(score_value=score_value)

            messages.success(request, "Formulario guardado.")
            return redirect("orbita_form_list")
        messages.error(request, "No se pudo guardar. Revisa los campos marcados e intenta nuevamente.")
        return render(request, "orbita/form_edit.html", {
            "orbita_form": orbita_form,
            "form": form,
            "formset": formset,
            "criteria_formset": criteria_formset,
            "orbita_client": client,
            "subscription": _get_or_create_subscription(request.user),
            "orbita_page": "formularios",
        })

    def _autosave_criteria(self, request, orbita_form):
        """
        Autosave de scores manuales (0-100) por AJAX sin requerir submit completo.
        Espera: criteria_payload='{"criteria":[{"id":1,"score":80}, ...]}'
        """
        raw_payload = (request.POST.get("criteria_payload") or "").strip()
        if not raw_payload:
            return JsonResponse({"ok": False, "error": "Sin datos de criterios."}, status=400)
        try:
            payload = json.loads(raw_payload)
        except (TypeError, ValueError, json.JSONDecodeError):
            return JsonResponse({"ok": False, "error": "Formato de payload inválido."}, status=400)

        items = payload.get("criteria")
        if not isinstance(items, list):
            return JsonResponse({"ok": False, "error": "El payload debe incluir una lista de criterios."}, status=400)

        updated = 0
        for item in items:
            if not isinstance(item, dict):
                continue
            criterion_id = item.get("id")
            score_value = item.get("score")
            try:
                criterion_id = int(criterion_id)
                score_value = int(score_value)
            except (TypeError, ValueError):
                continue
            score_value = max(0, min(100, score_value))
            updated += ATSFormCriterion.objects.filter(
                form=orbita_form,
                pk=criterion_id,
            ).update(score_value=score_value)

        return JsonResponse({"ok": True, "updated": updated})

    def _apply_posted_existing_criteria_scores(self, request, orbita_form):
        """
        Aplica scores de criterios existentes desde POST tradicional del formset
        sin bloquear el guardado completo si falta algún campo oculto.
        """
        for key, value in request.POST.items():
            if not key.startswith("criteria-") or not key.endswith("-id"):
                continue
            parts = key.split("-")
            if len(parts) < 3:
                continue
            idx = parts[1]
            criterion_id_raw = (value or "").strip()
            score_raw = (request.POST.get(f"criteria-{idx}-score_value") or "").strip()
            if not criterion_id_raw or not score_raw:
                continue
            try:
                criterion_id = int(criterion_id_raw)
                score_value = int(score_raw)
            except (TypeError, ValueError):
                continue
            score_value = max(0, min(100, score_value))
            ATSFormCriterion.objects.filter(
                form=orbita_form,
                pk=criterion_id,
            ).update(score_value=score_value)


class ATSFormDeleteView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """Eliminar formulario."""
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "forms"

    def post(self, request, pk):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        orbita_form = get_object_or_404(ATSForm, pk=pk, client=client)
        orbita_form.delete()
        messages.success(request, "Formulario eliminado.")
        return redirect("orbita_form_list")


class ATSFormSubmissionsView(OrbitaModuleRequiredMixin, LoginRequiredMixin, ListView):
    """Envíos recibidos de un formulario."""
    model = ATSFormSubmission
    template_name = "orbita/form_submissions.html"
    login_url = reverse_lazy("orbita_plataforma")
    context_object_name = "submissions"
    module_required = "forms"

    def get_queryset(self):
        client = _get_client_or_403(self.request)
        if not client:
            return ATSFormSubmission.objects.none()
        orbita_form = get_object_or_404(ATSForm, pk=self.kwargs["pk"], client=client)
        self.orbita_form = orbita_form
        submissions = list(
            ATSFormSubmission.objects.filter(form=orbita_form).prefetch_related("files").order_by("-submitted_at")
        )
        self._attach_ordered_payload_items(submissions, orbita_form)
        return submissions

    def _attach_ordered_payload_items(self, submissions, orbita_form):
        fields = list(orbita_form.fields.all().order_by("order", "id"))
        file_labels = {field.label for field in fields if field.field_type == ATSFormField.FIELD_FILE}
        file_labels.add("CV")
        ordered_labels = [field.label for field in fields if field.field_type != ATSFormField.FIELD_FILE]

        for submission in submissions:
            payload = submission.payload or {}
            items = []
            seen = set()
            for label in ordered_labels:
                if label in payload and label not in file_labels:
                    items.append({"label": label, "value": self._format_payload_value(payload.get(label))})
                    seen.add(label)
            for label, value in payload.items():
                if label in seen or label in file_labels:
                    continue
                items.append({"label": label, "value": self._format_payload_value(value)})
                seen.add(label)
            submission.display_payload_items = items

    def _format_payload_value(self, value):
        if isinstance(value, list):
            return ", ".join(str(v) for v in value if str(v).strip())
        return value

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["orbita_form"] = getattr(self, "orbita_form", None)
        context["orbita_client"] = _get_client_or_403(self.request)
        context["subscription"] = _get_or_create_subscription(self.request.user)
        context["orbita_page"] = "formularios"
        return context


class ATSFormSubmissionDeleteView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """POST: elimina un envío individual."""
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "forms"

    def post(self, request, pk, sub_pk):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        orbita_form = get_object_or_404(ATSForm, pk=pk, client=client)
        submission = get_object_or_404(ATSFormSubmission, pk=sub_pk, form=orbita_form)
        submission.delete()
        return redirect("orbita_form_submissions", pk=pk)


class ATSFormSubmissionDeleteAllView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """POST: elimina todos los envíos de un formulario."""
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "forms"

    def post(self, request, pk):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        orbita_form = get_object_or_404(ATSForm, pk=pk, client=client)
        ATSFormSubmission.objects.filter(form=orbita_form).delete()
        return redirect("orbita_form_submissions", pk=pk)


class ATSFormPublicView(View):
    """Formulario público: GET muestra el formulario, POST recibe el envío."""
    template_name = "orbita/form_public.html"
    thank_you_template = "orbita/form_public_thankyou.html"

    def _module_is_enabled(self, orbita_form):
        subscription = getattr(getattr(orbita_form.client, "user", None), "ats_subscription", None)
        return subscription_module_enabled(subscription, "forms")

    def get(self, request, uuid):
        orbita_form = get_object_or_404(ATSForm, uuid=uuid, is_active=True)
        if not self._module_is_enabled(orbita_form):
            return HttpResponseForbidden("Este formulario no está disponible.")
        has_email_field = orbita_form.fields.filter(field_type=ATSFormField.FIELD_EMAIL).exists()
        return render(request, self.template_name, {
            "orbita_form": orbita_form,
            "orbita_form_has_email_field": has_email_field,
        })

    def post(self, request, uuid):
        orbita_form = get_object_or_404(ATSForm, uuid=uuid, is_active=True)
        if not self._module_is_enabled(orbita_form):
            return HttpResponseForbidden("Este formulario no está disponible.")
        # Rate limit por IP + formulario
        ip = request.META.get("REMOTE_ADDR", "") or "unknown"
        cache_key = f"orbita_form_submit:{ip}:{uuid}"
        count = cache.get(cache_key, 0)
        max_count = getattr(settings, "ORBITA_FORM_PUBLIC_RATE_LIMIT_COUNT", 5)
        if count >= max_count:
            return render(request, self.template_name, {
                "orbita_form": orbita_form,
                "orbita_form_has_email_field": orbita_form.fields.filter(field_type=ATSFormField.FIELD_EMAIL).exists(),
                "form_error": "Has alcanzado el límite de envíos. Intenta de nuevo más tarde.",
            })
        # Validación de archivos (tamaño y extensión)
        max_size = getattr(settings, "ORBITA_FORM_PUBLIC_MAX_FILE_SIZE", 10 * 1024 * 1024)
        allowed_ext = getattr(settings, "ORBITA_FORM_PUBLIC_ALLOWED_EXTENSIONS", ["pdf", "doc", "docx"])
        def _check_file(f):
            if f.size > max_size:
                return False, f"El archivo {getattr(f, 'name', '')} supera el tamaño máximo permitido ({max_size // (1024*1024)} MB)."
            ext = (f.name or "").rsplit(".", 1)[-1].lower() if "." in (f.name or "") else ""
            if ext and allowed_ext and ext not in allowed_ext:
                return False, f"Solo se permiten archivos: {', '.join(allowed_ext)}."
            return True, None
        payload = {}
        files_to_save = []
        submitter_email = ""
        for field in orbita_form.fields.all().order_by("order", "id"):
            key = f"field_{field.id}"
            if field.field_type == ATSFormField.FIELD_FILE:
                f = request.FILES.get(key)
                if f and field.required or f:
                    ok, err = _check_file(f)
                    if not ok:
                        return render(request, self.template_name, {
                            "orbita_form": orbita_form,
                            "orbita_form_has_email_field": orbita_form.fields.filter(field_type=ATSFormField.FIELD_EMAIL).exists(),
                            "form_error": err,
                        })
                    files_to_save.append((field, f))
                    payload[field.label] = f.name
            elif field.field_type == ATSFormField.FIELD_MULTI:
                vals = [v.strip() for v in request.POST.getlist(key) if (v or "").strip()]
                allowed_options = {str(v).strip() for v in (field.option_values or []) if str(v).strip()}
                if allowed_options:
                    vals = [v for v in vals if v in allowed_options]
                if vals or field.required:
                    if field.required and not vals:
                        return render(request, self.template_name, {
                            "orbita_form": orbita_form,
                            "orbita_form_has_email_field": orbita_form.fields.filter(field_type=ATSFormField.FIELD_EMAIL).exists(),
                            "form_error": f"El campo «{field.label}» es obligatorio.",
                        })
                    payload[field.label] = vals
            else:
                val = request.POST.get(key, "").strip()
                if field.field_type == ATSFormField.FIELD_RADIO:
                    allowed_options = {str(v).strip() for v in (field.option_values or []) if str(v).strip()}
                    if allowed_options and val and val not in allowed_options:
                        val = ""
                if val or field.required:
                    if field.required and not val:
                        return render(request, self.template_name, {
                            "orbita_form": orbita_form,
                            "orbita_form_has_email_field": orbita_form.fields.filter(field_type=ATSFormField.FIELD_EMAIL).exists(),
                            "form_error": f"El campo «{field.label}» es obligatorio.",
                        })
                    payload[field.label] = val
                    if field.field_type == ATSFormField.FIELD_EMAIL and not submitter_email:
                        submitter_email = val
        if getattr(orbita_form, "request_cv", False):
            cv_file = request.FILES.get("cv_file")
            if cv_file:
                ok, err = _check_file(cv_file)
                if not ok:
                    return render(request, self.template_name, {
                        "orbita_form": orbita_form,
                        "orbita_form_has_email_field": orbita_form.fields.filter(field_type=ATSFormField.FIELD_EMAIL).exists(),
                        "form_error": err,
                    })
                files_to_save.append((None, cv_file))
                payload["CV"] = cv_file.name
        submitter_email = normalize_submitter_email(submitter_email or request.POST.get("submitter_email", "").strip())
        if submitter_email and "Correo electrónico" not in payload and "Email" not in payload:
            payload["Correo electrónico"] = submitter_email
        submission, duplicate_submission = create_submission_once(orbita_form, payload, submitter_email)
        if duplicate_submission:
            return render(request, self.thank_you_template, {
                "orbita_form": orbita_form,
                "already_submitted": True,
            })
        for field, uploaded_file in files_to_save:
            ATSFormSubmissionFile.objects.create(
                submission=submission,
                form_field=field,
                file=uploaded_file,
                original_name=uploaded_file.name,
            )
        # Si el formulario está ligado a una vacante, crear candidato y una sola notificación "Nuevo candidato"
        if orbita_form.vacancy_id:
            _create_candidate_from_submission(submission, payload, submitter_email)
            if submission.candidate_id:
                notify_orbita_client(
                    orbita_form.client,
                    ATSNotification.TYPE_CANDIDATE,
                    "Nuevo candidato",
                    message=f"{submission.candidate.name} — Formulario «{orbita_form.name}».",
                    link=reverse("orbita_candidate_detail", args=[submission.candidate.public_id]),
                    request=request,
                )
        else:
            # Sin vacante: notificación solo de envío
            notify_orbita_client(
                orbita_form.client,
                ATSNotification.TYPE_SUBMISSION,
                "Nuevo envío de formulario",
                message=f"Formulario «{orbita_form.name}»: {submitter_email or 'Sin correo'}.",
                link=reverse("orbita_form_submissions", args=[orbita_form.pk]),
                request=request,
            )
        # Rate limit: incrementar tras envío correcto
        timeout = getattr(settings, "ORBITA_FORM_PUBLIC_RATE_LIMIT_SECONDS", 3600)
        cache.set(cache_key, count + 1, timeout=timeout)
        return redirect("orbita_form_public_thanks", uuid=uuid)


class ATSFormPublicThanksView(View):
    """Página de agradecimiento tras enviar el formulario público (GET; se llega por redirect tras POST)."""
    template_name = "orbita/form_public_thankyou.html"

    def get(self, request, uuid):
        orbita_form = get_object_or_404(ATSForm, uuid=uuid, is_active=True)
        subscription = getattr(getattr(orbita_form.client, "user", None), "ats_subscription", None)
        if not subscription_module_enabled(subscription, "forms"):
            return HttpResponseForbidden("Este formulario no está disponible.")
        return render(request, self.template_name, {"orbita_form": orbita_form})


def _create_candidate_from_submission(submission, payload, submitter_email):
    """Crea un Candidato a partir de un envío de formulario y lo vincula. Usado en POST público y en backfill. Respeta límite de candidatos del plan."""
    orbita_form = submission.form
    if not orbita_form.vacancy_id:
        return None
    client = orbita_form.client
    subscription = getattr(client.user, "ats_subscription", None)
    current_count = Candidate.objects.filter(client=client).count()
    if not subscription_can_add_candidate(subscription, current_count):
        return None
    submitter_email = submitter_email or submission.submitter_email or ""
    name = (
        payload.get("Nombre")
        or payload.get("Nombre completo")
        or payload.get("Nombre y apellidos")
        or payload.get("name")
        or (submitter_email.split("@")[0] if submitter_email else "Postulante")
    )
    if not name or len(str(name)) > 255:
        name = (str(name) if name else "Postulante")[:255]
    candidate = Candidate.objects.create(
        client=orbita_form.client,
        vacancy=orbita_form.vacancy,
        name=name,
        email=submitter_email or "",
        status=Candidate.STATUS_REVISION,
        score=0,
        explanation_text="Postulación recibida por formulario. Pendiente de análisis de CV con IA.",
    )
    submission.candidate = candidate
    submission.save(update_fields=["candidate"])
    # Si el envío incluyó un archivo (CV), copiarlo al candidato para poder procesarlo con IA después
    # Preferir el archivo del campo "Solicitar CV" (form_field=None); si no, el primer archivo adjunto
    cv_attachment = submission.files.filter(form_field__isnull=True).first() or submission.files.first()
    if cv_attachment and cv_attachment.file:
        from django.core.files.base import ContentFile
        try:
            name = cv_attachment.original_name or "cv_adjunto"
            if not name or "." not in name:
                name = name or "cv_adjunto"
            candidate.cv_file.save(name, ContentFile(cv_attachment.file.read()), save=True)
        except Exception:
            pass

    _auto_analyze_candidate_if_applicable(candidate)
    return candidate


def _auto_analyze_candidate_if_applicable(candidate):
    """
    Ejecuta análisis automático de CV para postulaciones nuevas cuando aplica:
    - Config global de análisis CV activa.
    - IA activa en la vacante (si hay vacante).
    - Candidato con CV adjunto.
    - Plan con capacidad y cupo de análisis disponible.
    """
    if not candidate:
        return False

    try:
        cv_config, _ = CVAnalysisConfig.objects.get_or_create(client=candidate.client)
        if not cv_config.enabled:
            return False

        if candidate.vacancy_id and not getattr(candidate.vacancy, "ai_enabled", False):
            return False

        if not candidate.cv_file:
            return False

        subscription = _get_or_create_subscription(candidate.client.user)
        if not subscription_can(subscription, "cvs_scan"):
            return False

        from mi_app.services.cv_analysis import run_cv_analysis_and_save
        result = run_cv_analysis_and_save(candidate)
        if not result.get("ok"):
            logger.warning(
                "Auto análisis CV no aplicado para candidate=%s: %s",
                candidate.pk,
                result.get("error", "error desconocido"),
            )
            return False

        subscription.increment_cvs_used()
        return True
    except Exception as exc:
        logger.warning("Error en auto análisis CV candidate=%s: %s", getattr(candidate, "pk", None), exc)
        return False


class ATSEmailConfigView(OrbitaModuleRequiredMixin, LoginRequiredMixin, FormView):
    """Configuración de correo: notificaciones y correo de la empresa (conectar inbox)."""
    template_name = "orbita/email_config.html"
    form_class = ATSEmailConfigForm
    login_url = reverse_lazy("orbita_plataforma")
    success_url = reverse_lazy("orbita_email_config")
    module_required = "email_config"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        client = _get_client_or_403(self.request)
        if not client:
            return kwargs
        config, _ = ATSClientEmailConfig.objects.get_or_create(client=client)
        kwargs["instance"] = config
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["orbita_client"] = _get_client_or_403(self.request)
        context["subscription"] = _get_or_create_subscription(self.request.user)
        context["orbita_page"] = "correo"
        return context

    def form_valid(self, form):
        client = _get_client_or_403(self.request)
        if not client:
            return redirect("orbita_dashboard")
        form.instance.client = client
        form.save()
        smtp_password_changed = bool((form.cleaned_data.get("smtp_password") or "").strip())
        imap_password_changed = bool((form.cleaned_data.get("imap_password") or "").strip())
        if smtp_password_changed or imap_password_changed:
            messages.success(self.request, "Configuración de correo guardada. Las contraseñas nuevas fueron actualizadas.")
        else:
            messages.success(self.request, "Configuración de correo guardada.")
        return super().form_valid(form)

    def post(self, request, *args, **kwargs):
        if request.POST.get("test_smtp"):
            return self._handle_test_smtp(request)
        return super().post(request, *args, **kwargs)

    def _handle_test_smtp(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")

        config, _ = ATSClientEmailConfig.objects.get_or_create(client=client)
        form = self.form_class(request.POST, instance=config)
        if not form.is_valid():
            messages.error(request, "Revisa los datos SMTP e intenta nuevamente.")
            return render(request, self.template_name, self.get_context_data(form=form))

        smtp_host = (form.cleaned_data.get("smtp_host") or "").strip()
        smtp_user = (form.cleaned_data.get("smtp_user") or "").strip()
        smtp_port = int(form.cleaned_data.get("smtp_port") or 587)
        smtp_use_tls = bool(form.cleaned_data.get("smtp_use_tls"))
        smtp_password = (form.cleaned_data.get("smtp_password") or "").strip() or (
            getattr(config, "smtp_password_encrypted", "") or ""
        ).strip()

        if not smtp_host or not smtp_user or not smtp_password:
            messages.error(
                request,
                "Para probar SMTP completa servidor, usuario y contraseña en Config. correo.",
            )
            return render(request, self.template_name, self.get_context_data(form=form))

        from_name = (form.cleaned_data.get("company_from_name") or "").strip() or client.company_name or "ATS"
        from_email = (form.cleaned_data.get("company_from_email") or "").strip() or smtp_user
        from_header = f"{from_name} <{from_email}>"
        notification_email = (form.cleaned_data.get("notification_email") or "").strip()
        to_email = notification_email or (request.user.email or "").strip() or smtp_user

        connection = EmailBackend(
            host=smtp_host,
            port=smtp_port,
            username=smtp_user,
            password=smtp_password,
            use_tls=smtp_use_tls,
            timeout=getattr(settings, "EMAIL_TIMEOUT", 10),
            fail_silently=False,
        )
        try:
            connection.open()
            send_mail(
                subject="[ATS] Prueba SMTP",
                message=(
                    "Prueba SMTP exitosa.\n\n"
                    "Este correo confirma que la configuración SMTP del cliente funciona correctamente."
                ),
                from_email=from_header,
                recipient_list=[to_email],
                fail_silently=False,
                connection=connection,
            )
            messages.success(
                request,
                f"SMTP verificado correctamente. Se envió un correo de prueba a {to_email}.",
            )
        except Exception as exc:
            error_text = str(exc)
            lower_error = error_text.lower()
            if ("5.7.9" in lower_error) or ("application-specific password required" in lower_error):
                messages.error(
                    request,
                    "Google bloqueó el acceso SMTP. Usa una contraseña de aplicación (App Password) en vez de la contraseña normal.",
                )
            else:
                messages.error(request, f"No se pudo validar SMTP: {exc}")
        finally:
            try:
                connection.close()
            except Exception:
                pass

        return render(request, self.template_name, self.get_context_data(form=form))


class ATSProfileConfigView(OrbitaModuleRequiredMixin, LoginRequiredMixin, FormView):
    """Configurar cuenta: foto de perfil, nombre, teléfono, empresa."""
    template_name = "orbita/profile_config.html"
    form_class = ATSProfileForm
    login_url = reverse_lazy("orbita_plataforma")
    success_url = reverse_lazy("orbita_profile_config")
    module_required = "account"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        client = _get_client_or_403(self.request)
        if client:
            kwargs["instance"] = client
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["orbita_client"] = _get_client_or_403(self.request)
        subscription = _get_or_create_subscription(self.request.user)
        context["subscription"] = subscription
        context["kpi_cvs_used"] = subscription.cvs_used
        context["kpi_cvs_limit"] = subscription.cvs_limit
        context["subscription_active"] = subscription.active
        context["orbita_page"] = "perfil"
        return context

    def form_valid(self, form):
        client = _get_client_or_403(self.request)
        if not client:
            return redirect("orbita_dashboard")
        form.save()
        messages.success(self.request, "Cuenta actualizada.")
        return redirect(self.get_success_url())


class ATSProfileAvatarUploadView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """Actualiza solo la foto de perfil desde Configurar cuenta."""
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "account"
    http_method_names = ["post"]

    def post(self, request):
        client = _get_client_or_403(request)
        if not client:
            return JsonResponse({"ok": False, "error": "Cuenta no encontrada."}, status=404)
        form = ATSAvatarUploadForm(request.POST, request.FILES, instance=client)
        if not form.is_valid():
            avatar_errors = form.errors.get("avatar")
            error = avatar_errors[0] if avatar_errors else "No se pudo actualizar la imagen."
            return JsonResponse({"ok": False, "error": str(error)}, status=400)
        form.save()
        return JsonResponse({
            "ok": True,
            "avatar_url": client.avatar.url if client.avatar else "",
            "message": "Foto actualizada.",
        })


class ATSVacancyCreateView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """Crear vacante (puesto) desde Reclutamiento."""
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "vacancies"

    @method_decorator(never_cache)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        subscription = _get_or_create_subscription(request.user)
        vacancy_count = Vacancy.objects.filter(client=client).count()
        if not subscription_can_add_vacancy(subscription, vacancy_count):
            messages.error(request, "Has alcanzado el límite de vacantes de tu plan (Gratuito: 2 vacantes). Mejora tu plan para crear más.")
            return redirect(reverse("orbita_dashboard") + "?section=reclutamiento")
        return render(request, "orbita/vacancy_form.html", {
            "form": ATSVacancyForm(),
            "orbita_client": client,
            "subscription": subscription,
            "orbita_page": "reclutamiento",
        })

    def post(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        subscription = _get_or_create_subscription(request.user)
        vacancy_count = Vacancy.objects.filter(client=client).count()
        if not subscription_can_add_vacancy(subscription, vacancy_count):
            messages.error(request, "Has alcanzado el límite de vacantes de tu plan. Mejora tu plan para crear más.")
            return redirect(reverse("orbita_dashboard") + "?section=reclutamiento")
        form = ATSVacancyForm(request.POST)
        if form.is_valid():
            vacancy = form.save(commit=False)
            vacancy.client = client
            vacancy.save()
            messages.success(request, "Vacante creada. Ya puedes asociar candidatos y formularios a este puesto.")
            return redirect(reverse("orbita_dashboard") + "?section=reclutamiento")
        return render(request, "orbita/vacancy_form.html", {
            "form": form,
            "orbita_client": client,
            "subscription": subscription,
            "orbita_page": "reclutamiento",
        })


class ATSVacancyEditView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """Editar vacante (título, descripción, perfil para análisis con IA)."""
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "vacancies"

    def get(self, request, public_id):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        vacancy = get_object_or_404(Vacancy, public_id=public_id, client=client)
        form = ATSVacancyForm(instance=vacancy)
        return render(request, "orbita/vacancy_form.html", {
            "form": form,
            "orbita_client": client,
            "subscription": _get_or_create_subscription(request.user),
            "orbita_page": "reclutamiento",
            "vacancy": vacancy,
            "is_edit": True,
        })

    def post(self, request, public_id):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        vacancy = get_object_or_404(Vacancy, public_id=public_id, client=client)
        form = ATSVacancyForm(request.POST, instance=vacancy)
        if form.is_valid():
            form.save()
            messages.success(request, "Vacante actualizada.")
            return redirect(reverse("orbita_dashboard") + "?section=reclutamiento")
        return render(request, "orbita/vacancy_form.html", {
            "form": form,
            "orbita_client": client,
            "subscription": _get_or_create_subscription(request.user),
            "orbita_page": "reclutamiento",
            "vacancy": vacancy,
            "is_edit": True,
        })


class ATSVacancyDeleteView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """Eliminar vacante."""
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "vacancies"

    def post(self, request, public_id):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        vacancy = get_object_or_404(Vacancy, public_id=public_id, client=client)
        vacancy.delete()
        messages.success(request, "Vacante eliminada.")
        return redirect(reverse("orbita_dashboard") + "?section=reclutamiento")


def _redirect_workforce(tab="necesidad"):
    return redirect(f"{reverse('orbita_workforce_dashboard')}?tab={tab}")


class ATSWorkforceDashboardView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """Dashboard Workforce: áreas, puestos, necesidades, aprobaciones y brechas."""
    template_name = "orbita/workforce_dashboard.html"
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "workforce"

    def get(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        valid_tabs = {"necesidad", "areas", "puestos", "planeacion"}
        active_tab = request.GET.get("tab")
        if active_tab in {"resumen", "planes"}:
            active_tab = "necesidad"
        if active_tab not in valid_tabs:
            active_tab = "necesidad"
        areas = WorkforceArea.objects.filter(client=client).prefetch_related("positions")
        positions = WorkforcePosition.objects.filter(client=client).select_related("area")
        plans = WorkforcePlan.objects.filter(client=client).select_related("area", "position")[:100]
        edit_plan = None
        edit_public_id = request.GET.get("edit")
        if edit_public_id:
            try:
                edit_uuid = uuid_lib.UUID(str(edit_public_id))
                edit_plan = WorkforcePlan.objects.filter(client=client, public_id=edit_uuid).select_related("area", "position").first()
                if edit_plan:
                    active_tab = "planeacion"
            except (TypeError, ValueError):
                edit_plan = None
        total_current = sum(plan.current_staff for plan in plans)
        total_required = sum(plan.required_staff for plan in plans)
        total_gap = sum(plan.gap for plan in plans)
        total_budget = sum(plan.estimated_budget for plan in plans)
        pending_count = sum(1 for plan in plans if plan.status == WorkforcePlan.STATUS_PENDING)
        approved_count = sum(1 for plan in plans if plan.status == WorkforcePlan.STATUS_APPROVED)
        high_risk_count = sum(1 for plan in plans if plan.operational_risk == "Alto")
        return render(request, self.template_name, {
            "orbita_client": client,
            "subscription": _get_or_create_subscription(request.user),
            "orbita_page": "workforce",
            "active_tab": active_tab,
            "areas": areas,
            "positions": positions,
            "plans": plans,
            "edit_plan": edit_plan,
            "area_form": WorkforceAreaForm(),
            "position_form": WorkforcePositionForm(client=client),
            "plan_form": WorkforcePlanForm(client=client),
            "total_current": total_current,
            "total_required": total_required,
            "total_gap": total_gap,
            "total_budget": total_budget,
            "pending_count": pending_count,
            "approved_count": approved_count,
            "high_risk_count": high_risk_count,
        })


class ATSWorkforceAreaCreateView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "workforce"
    http_method_names = ["post"]

    def post(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        form = WorkforceAreaForm(request.POST)
        if form.is_valid():
            area = form.save(commit=False)
            area.client = client
            area.save()
            messages.success(request, "Área agregada a Workforce.")
        else:
            messages.error(request, "No se pudo crear el área. Revisa el nombre.")
        return _redirect_workforce("areas")


class ATSWorkforceAreaUpdateView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "workforce"
    http_method_names = ["post"]

    def post(self, request, public_id):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        area = get_object_or_404(WorkforceArea, public_id=public_id, client=client)
        form = WorkforceAreaForm(request.POST, instance=area)
        if form.is_valid():
            form.save()
            messages.success(request, "Área actualizada.")
        else:
            messages.error(request, "No se pudo actualizar el área.")
        return _redirect_workforce("areas")


class ATSWorkforceAreaDeleteView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "workforce"
    http_method_names = ["post"]

    def post(self, request, public_id):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        area = get_object_or_404(WorkforceArea, public_id=public_id, client=client)
        if area.plans.exists():
            messages.error(request, "No puedes eliminar un área con planes Workforce asociados.")
            return _redirect_workforce("areas")
        area.delete()
        messages.success(request, "Área eliminada.")
        return _redirect_workforce("areas")


class ATSWorkforcePositionCreateView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "workforce"
    http_method_names = ["post"]

    def post(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        form = WorkforcePositionForm(request.POST, client=client)
        if form.is_valid():
            position = form.save(commit=False)
            position.client = client
            position.save()
            messages.success(request, "Puesto agregado a Workforce.")
        else:
            messages.error(request, "No se pudo crear el puesto. Revisa área y salarios.")
        return _redirect_workforce("puestos")


class ATSWorkforcePositionUpdateView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "workforce"
    http_method_names = ["post"]

    def post(self, request, public_id):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        position = get_object_or_404(WorkforcePosition, public_id=public_id, client=client)
        form = WorkforcePositionForm(request.POST, client=client, instance=position)
        if form.is_valid():
            position = form.save()
            for plan in position.plans.all():
                plan.save(update_fields=["estimated_budget", "updated_at"])
            messages.success(request, "Puesto actualizado.")
        else:
            messages.error(request, "No se pudo actualizar el puesto.")
        return _redirect_workforce("puestos")


class ATSWorkforcePositionDeleteView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "workforce"
    http_method_names = ["post"]

    def post(self, request, public_id):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        position = get_object_or_404(WorkforcePosition, public_id=public_id, client=client)
        if position.plans.exists():
            messages.error(request, "No puedes eliminar un puesto con planes Workforce asociados.")
            return _redirect_workforce("puestos")
        position.delete()
        messages.success(request, "Puesto eliminado.")
        return _redirect_workforce("puestos")


class ATSWorkforcePlanCreateView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "workforce"
    http_method_names = ["post"]

    def post(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        form = WorkforcePlanForm(request.POST, client=client)
        if form.is_valid():
            plan = form.save(commit=False)
            plan.client = client
            plan.save()
            messages.success(request, "Necesidad de personal registrada.")
        else:
            messages.error(request, "No se pudo registrar la necesidad. Revisa área, puesto y cantidades.")
        return _redirect_workforce("necesidad")


class ATSWorkforcePlanUpdateView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "workforce"
    http_method_names = ["post"]

    def post(self, request, public_id):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        plan = get_object_or_404(WorkforcePlan, public_id=public_id, client=client)
        if plan.status == WorkforcePlan.STATUS_CONVERTED:
            messages.error(request, "No puedes editar una necesidad ya convertida en vacante.")
            return _redirect_workforce("necesidad")
        form = WorkforcePlanForm(request.POST, client=client, instance=plan)
        if form.is_valid():
            form.save()
            messages.success(request, "Necesidad actualizada.")
        else:
            messages.error(request, "No se pudo actualizar la necesidad.")
        return _redirect_workforce("necesidad")


class ATSWorkforcePlanDeleteView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "workforce"
    http_method_names = ["post"]

    def post(self, request, public_id):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        plan = get_object_or_404(WorkforcePlan, public_id=public_id, client=client)
        if plan.created_vacancies.exists():
            messages.error(request, "No puedes eliminar una necesidad que ya creó una vacante.")
            return _redirect_workforce("necesidad")
        plan.delete()
        messages.success(request, "Necesidad eliminada.")
        return _redirect_workforce("necesidad")


class ATSWorkforcePlanActionView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "workforce"
    http_method_names = ["post"]
    action = None

    def post(self, request, public_id):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        plan = get_object_or_404(WorkforcePlan, public_id=public_id, client=client)
        if self.action == "approve":
            plan.status = WorkforcePlan.STATUS_APPROVED
            plan.save(update_fields=["status", "estimated_budget", "updated_at"])
            messages.success(request, "Necesidad aprobada.")
        elif self.action == "reject":
            plan.status = WorkforcePlan.STATUS_REJECTED
            plan.save(update_fields=["status", "estimated_budget", "updated_at"])
            messages.success(request, "Necesidad rechazada.")
        elif self.action == "convert":
            return self._convert_to_vacancy(request, plan)
        return _redirect_workforce("necesidad")

    def _convert_to_vacancy(self, request, plan):
        if plan.status == WorkforcePlan.STATUS_CONVERTED:
            messages.info(request, "Esta necesidad ya fue convertida en vacante.")
            return _redirect_workforce("necesidad")
        if plan.status != WorkforcePlan.STATUS_APPROVED:
            messages.error(request, "Solo puedes convertir necesidades aprobadas.")
            return _redirect_workforce("necesidad")
        if plan.gap <= 0:
            messages.error(request, "La necesidad no tiene brecha de personal para convertir.")
            return _redirect_workforce("necesidad")

        subscription = _get_or_create_subscription(request.user)
        vacancy_count = Vacancy.objects.filter(client=plan.client).count()
        if not subscription_can_add_vacancy(subscription, vacancy_count):
            messages.error(request, "Has alcanzado el límite de vacantes de tu plan.")
            return _redirect_workforce("necesidad")

        Vacancy.objects.create(
            client=plan.client,
            title=plan.position.name,
            description=(
                f"Vacante generada desde Workforce.\n\n"
                f"Área: {plan.area.name}\n"
                f"Plazas requeridas: {plan.gap}\n"
                f"Prioridad: {plan.get_priority_display()}\n"
                f"Riesgo operativo: {plan.operational_risk}\n\n"
                f"{plan.executive_justification}".strip()
            ),
            profile_for_analysis=(
                f"Evaluar compatibilidad para el puesto {plan.position.name} en el área {plan.area.name}. "
                f"Prioridad {plan.get_priority_display()} y brecha de {plan.gap} persona(s)."
            ),
            openings=plan.gap,
            salary_min=plan.position.salary_min,
            salary_max=plan.position.salary_max,
            ai_enabled=True,
            source=Vacancy.SOURCE_WORKFORCE,
            workforce_plan=plan,
        )
        plan.status = WorkforcePlan.STATUS_CONVERTED
        plan.save(update_fields=["status", "estimated_budget", "updated_at"])
        messages.success(request, "Vacante creada desde Workforce. Ya aparece en Vacantes.")
        return redirect(reverse("orbita_dashboard") + "?section=reclutamiento")


class ATSCVAnalysisConfigView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """Configuración por defecto del análisis de CV con IA (perfil e instrucciones)."""
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "cv_analysis"

    def get(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        config, _ = CVAnalysisConfig.objects.get_or_create(client=client)
        form = CVAnalysisConfigForm(instance=config)
        return render(request, "orbita/cv_analysis_config.html", {
            "form": form,
            "orbita_client": client,
            "subscription": _get_or_create_subscription(request.user),
            "orbita_page": "analisis_cv",
        })

    def post(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        config, _ = CVAnalysisConfig.objects.get_or_create(client=client)
        form = CVAnalysisConfigForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            messages.success(request, "Configuración de análisis de CV guardada.")
            return redirect("orbita_cv_analysis_config")
        return render(request, "orbita/cv_analysis_config.html", {
            "form": form,
            "orbita_client": client,
            "subscription": _get_or_create_subscription(request.user),
            "orbita_page": "analisis_cv",
        })


class ATSBackfillCandidatesFromSubmissionsView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """Crea Candidatos desde envíos de formularios que tienen vacante pero aún no tienen candidato vinculado."""
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "vacancies"
    http_method_names = ["post"]

    def post(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        pending = ATSFormSubmission.objects.filter(
            form__client=client,
            form__vacancy__isnull=False,
            candidate__isnull=True,
        ).select_related("form", "form__vacancy")
        created = 0
        for sub in pending:
            if _create_candidate_from_submission(sub, sub.payload, sub.submitter_email):
                created += 1
        if created:
            messages.success(
                request,
                f"Se crearon {created} candidato(s) desde envíos de formularios. Ya aparecen en Vacantes y Candidatos.",
            )
            notify_orbita_client(
                client,
                ATSNotification.TYPE_CANDIDATE,
                "Candidatos creados desde envíos",
                message=f"Se crearon {created} candidato(s) desde formularios. Revisa la sección Candidatos.",
                link=reverse("orbita_dashboard") + "?section=candidatos",
                request=request,
            )
        else:
            messages.info(request, "No había envíos pendientes de vincular. Todos los envíos con vacante ya tienen candidato.")
        return redirect(reverse("orbita_dashboard") + "?section=reclutamiento")


class ATSNotificationGoView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """Marca una notificación como leída y redirige a su enlace."""
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "notifications"
    http_method_names = ["get"]

    def get(self, request, pk):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        notification = get_object_or_404(ATSNotification, pk=pk, client=client)
        notification.read = True
        notification.save(update_fields=["read"])
        link = (notification.link or "").strip()
        if link and "/candidato/" in link:
            try:
                cand_pk = int(link.rstrip("/").split("/")[-1])
                Candidate.objects.get(pk=cand_pk, client=client)
            except (Candidate.DoesNotExist, ValueError, IndexError):
                messages.warning(request, "El candidato vinculado a esta notificación ya no existe.")
                return redirect("orbita_dashboard")
        if link and not link.startswith("http"):
            return redirect(link)
        return redirect(link or reverse("orbita_dashboard"))


class ATSNotificationMarkAllReadView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """Marca todas las notificaciones del cliente como leídas."""
    login_url = reverse_lazy("orbita_plataforma")
    module_required = "notifications"
    http_method_names = ["post"]

    def post(self, request):
        client = _get_client_or_403(request)
        if client:
            ATSNotification.objects.filter(client=client, read=False).update(read=True)
        ref = request.META.get("HTTP_REFERER") or reverse("orbita_dashboard")
        return redirect(ref)


class ATSNotificationPanelView(OrbitaModuleRequiredMixin, LoginRequiredMixin, View):
    """Panel de notificaciones: lista paginada con opción de marcar todas como leídas."""
    login_url = reverse_lazy("orbita_plataforma")
    template_name = "orbita/notification_panel.html"
    paginate_by = 25
    module_required = "notifications"

    def get(self, request):
        client = _get_client_or_403(request)
        if not client:
            return redirect("orbita_dashboard")
        qs = ATSNotification.objects.filter(client=client).order_by("-created_at")
        paginator = Paginator(qs, self.paginate_by)
        page_number = request.GET.get("page", 1)
        page = paginator.get_page(page_number)
        unread_count = ATSNotification.objects.filter(client=client, read=False).count()
        return render(request, self.template_name, {
            "page_obj": page,
            "notifications": page.object_list,
            "orbita_client": client,
            "orbita_unread_count": unread_count,
            "orbita_page": "notificaciones",
            "subscription": _get_or_create_subscription(request.user),
        })

    def post(self, request):
        """Marcar todas como leídas y redirigir al panel."""
        client = _get_client_or_403(request)
        if client:
            ATSNotification.objects.filter(client=client, read=False).update(read=True)
        return redirect("orbita_notification_panel")


# --- Panel de administración ATS (solo staff) ---

class ATSAdminDashboardView(StaffRequiredMixin, View):
    """Panel para soporte: listar clientes ATS y sus suscripciones; cambiar plan."""
    template_name = "orbita/admin/dashboard.html"
    paginate_by = 25

    def get(self, request):
        qs = ATSClient.objects.select_related("user").annotate(
            num_vacancies=Count("vacancies", distinct=True),
            num_candidates=Count("candidates", distinct=True),
            num_forms=Count("ats_forms", distinct=True),
        ).order_by("-created_at")
        search_q = (request.GET.get("q") or "").strip()
        if search_q:
            qs = qs.filter(
                Q(company_name__icontains=search_q)
                | Q(contact_name__icontains=search_q)
                | Q(user__email__icontains=search_q)
            )
        paginator = Paginator(qs, self.paginate_by)
        page_number = request.GET.get("page", 1)
        page_obj = paginator.get_page(page_number)
        client_ids = [c.user_id for c in page_obj.object_list]
        subs_by_user = {
            sub.user_id: sub
            for sub in Subscription.objects.filter(user_id__in=client_ids).select_related("user")
        }
        rows = []
        for client in page_obj.object_list:
            rows.append({
                "client": client,
                "subscription": subs_by_user.get(client.user_id),
                "num_vacancies": client.num_vacancies,
                "num_candidates": client.num_candidates,
                "num_forms": client.num_forms,
                "last_login": client.user.last_login,
            })
        plan_choices = [
            (Subscription.PLAN_FREE, "Gratuito"),
            (Subscription.PLAN_PRO, "Pro"),
            (Subscription.PLAN_ENTERPRISE, "Enterprise"),
        ]
        pending_requests = list(
            PlanChangeRequest.objects.filter(status=PlanChangeRequest.STATUS_PENDING)
            .select_related("client")
            .order_by("-created_at")[:20]
        )
        # Uso de IA (tokens): totales para mostrar en admin y enlace a LangSmith
        now = timezone.now()
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        llm_usage_this_month = LLMUsageLog.objects.filter(created_at__gte=month_start).aggregate(
            total=Sum("total_tokens"), runs=Count("id")
        )
        llm_usage_all_time = LLMUsageLog.objects.aggregate(total=Sum("total_tokens"), runs=Count("id"))
        import os
        langsmith_configured = bool(os.environ.get("LANGSMITH_API_KEY", "").strip())
        langsmith_project = os.environ.get("LANGSMITH_PROJECT", "default")
        langsmith_runs = []
        if langsmith_configured:
            try:
                from langsmith import Client
                client = Client()
                for run in client.list_runs(project_name=langsmith_project, limit=15):
                    extra = getattr(run, "extra", None) or {}
                    usage = extra.get("usage") or {}
                    langsmith_runs.append({
                        "name": getattr(run, "name", "—"),
                        "start_time": getattr(run, "start_time", None),
                        "total_tokens": usage.get("total_tokens"),
                        "inputs": getattr(run, "inputs", None),
                    })
            except Exception:
                langsmith_runs = []
        total_clients = ATSClient.objects.count()
        total_pro = Subscription.objects.filter(plan=Subscription.PLAN_PRO).count()
        total_enterprise = Subscription.objects.filter(plan=Subscription.PLAN_ENTERPRISE).count()
        total_pending = len(pending_requests)
        total_vacancies = Vacancy.objects.count()
        total_candidates = Candidate.objects.count()
        total_forms_global = ATSForm.objects.count()
        top_clients = list(
            ATSClient.objects.select_related("user")
            .annotate(
                n_vac=Count("vacancies", distinct=True),
                n_cand=Count("candidates", distinct=True),
                n_forms=Count("ats_forms", distinct=True),
            )
            .order_by("-n_cand")[:5]
        )
        recent_clients = list(
            ATSClient.objects.select_related("user")
            .order_by("-created_at")[:5]
        )
        all_clients_for_notify = list(
            ATSClient.objects.order_by("company_name").values("id", "company_name")
        )
        return render(request, self.template_name, {
            "rows": rows,
            "page_obj": page_obj,
            "search_q": search_q,
            "plan_choices": plan_choices,
            "pending_requests": pending_requests,
            "llm_usage_this_month": llm_usage_this_month.get("total") or 0,
            "llm_usage_this_month_runs": llm_usage_this_month.get("runs") or 0,
            "llm_usage_all_time": llm_usage_all_time.get("total") or 0,
            "llm_usage_all_time_runs": llm_usage_all_time.get("runs") or 0,
            "langsmith_configured": langsmith_configured,
            "langsmith_project": langsmith_project,
            "langsmith_runs": langsmith_runs,
            "total_clients": total_clients,
            "total_pro": total_pro,
            "total_enterprise": total_enterprise,
            "total_pending": total_pending,
            "total_vacancies": total_vacancies,
            "total_candidates": total_candidates,
            "total_forms_global": total_forms_global,
            "top_clients": top_clients,
            "recent_clients": recent_clients,
            "all_clients_for_notify": all_clients_for_notify,
            "orbita_page": "administracion",
            "orbita_client": getattr(request.user, "ats_client", None),
            "orbita_header_title": "Administración ATS",
        })


class ATSAdminChangePlanView(StaffRequiredMixin, View):
    """POST: aplicar un plan a una suscripción (tras validar pago)."""
    def post(self, request):
        subscription_id = request.POST.get("subscription_id")
        plan_id = (request.POST.get("plan") or "").strip().upper()
        if not subscription_id or plan_id not in (Subscription.PLAN_FREE, Subscription.PLAN_PRO, Subscription.PLAN_ENTERPRISE):
            messages.error(request, "Datos inválidos.")
            return redirect("orbita_admin_dashboard")
        subscription = get_object_or_404(Subscription, pk=subscription_id)
        old_plan = subscription.plan
        ok = apply_plan_to_subscription(subscription, plan_id)
        if not ok:
            messages.error(request, "No se pudo aplicar el plan.")
            return redirect("orbita_admin_dashboard")
        messages.success(request, f"Plan actualizado a {subscription.get_plan_display()} para {subscription.user.email}.")
        client = getattr(subscription.user, "ats_client", None)
        if client:
            PlanChangeRequest.objects.filter(
                client=client, to_plan=plan_id, status=PlanChangeRequest.STATUS_PENDING
            ).update(status=PlanChangeRequest.STATUS_DONE)
            notify_orbita_client(
                client,
                ATSNotification.TYPE_PLAN,
                "Plan actualizado",
                message=f"Tu plan ha sido actualizado a {subscription.get_plan_display()}.",
                link=reverse("orbita_profile_config"),
                request=request,
            )
        return redirect("orbita_admin_dashboard")


class ATSAdminUpdateModulesView(StaffRequiredMixin, View):
    """POST: activar/desactivar módulos visibles para un cliente."""
    http_method_names = ["post"]

    MODULE_FIELDS = (
        "module_candidates",
        "module_vacancies",
        "module_forms",
        "module_account",
        "module_notifications",
        "module_email_config",
        "module_cv_analysis",
        "module_workforce",
    )

    def post(self, request):
        subscription_id = request.POST.get("subscription_id")
        if not subscription_id:
            messages.error(request, "Falta la suscripción.")
            return redirect("orbita_admin_dashboard")
        subscription = get_object_or_404(Subscription, pk=subscription_id)
        for field in self.MODULE_FIELDS:
            setattr(subscription, field, request.POST.get(field) == "on")
        subscription.save(update_fields=[*self.MODULE_FIELDS, "updated_at"])
        messages.success(request, f"Módulos actualizados para {subscription.user.email}.")
        return redirect("orbita_admin_dashboard")


class ATSAdminUpdateCVLimitView(StaffRequiredMixin, View):
    """POST: ajustar manualmente el límite mensual de análisis CV de una suscripción."""
    http_method_names = ["post"]

    def post(self, request):
        subscription_id = request.POST.get("subscription_id")
        cvs_limit_raw = (request.POST.get("cvs_limit") or "").strip()
        if not subscription_id:
            messages.error(request, "Falta la suscripción.")
            return redirect("orbita_admin_dashboard")
        try:
            cvs_limit = int(cvs_limit_raw)
        except (TypeError, ValueError):
            messages.error(request, "El límite de CVs debe ser un número entero.")
            return redirect("orbita_admin_dashboard")
        if cvs_limit < 0:
            messages.error(request, "El límite de CVs no puede ser negativo.")
            return redirect("orbita_admin_dashboard")
        if cvs_limit > 100000:
            messages.error(request, "El límite de CVs es demasiado alto.")
            return redirect("orbita_admin_dashboard")

        subscription = get_object_or_404(Subscription, pk=subscription_id)
        subscription.cvs_limit = cvs_limit
        subscription.save(update_fields=["cvs_limit", "updated_at"])
        messages.success(request, f"Límite de CVs actualizado a {cvs_limit} para {subscription.user.email}.")
        return redirect("orbita_admin_dashboard")


class ATSAdminSetLangSmithView(StaffRequiredMixin, View):
    """POST: asignar proyecto LangSmith a un cliente (desde el panel ATS administración)."""
    http_method_names = ["post"]

    def post(self, request):
        client_id = request.POST.get("client_id")
        langsmith_project = (request.POST.get("langsmith_project") or "").strip()[:100]
        if not client_id:
            messages.error(request, "Falta el cliente.")
            return redirect("orbita_admin_dashboard")
        client = get_object_or_404(ATSClient, pk=client_id)
        client.langsmith_project = langsmith_project
        client.save(update_fields=["langsmith_project"])
        if langsmith_project:
            messages.success(request, f"Proyecto LangSmith «{langsmith_project}» asignado a {client.company_name}.")
        else:
            messages.success(request, f"Proyecto LangSmith quitado para {client.company_name}. Se usará el global.")
        return redirect("orbita_admin_dashboard")


class ATSAdminDeleteClientView(StaffRequiredMixin, View):
    """POST: eliminar un cliente ATS y su usuario asociado."""
    http_method_names = ["post"]

    def post(self, request):
        client_id = request.POST.get("client_id")
        if not client_id:
            messages.error(request, "Falta el identificador del cliente.")
            return redirect("orbita_admin_dashboard")
        client = get_object_or_404(ATSClient, pk=client_id)
        company = client.company_name
        user = client.user
        Subscription.objects.filter(user=user).delete()
        client.delete()
        user.delete()
        messages.success(request, f"Cliente «{company}» y su cuenta eliminados correctamente.")
        return redirect("orbita_admin_dashboard")


class ATSAdminSendNotificationView(StaffRequiredMixin, View):
    """POST: enviar notificación desde admin a uno o todos los clientes."""
    http_method_names = ["post"]

    def post(self, request):
        from mi_app.orbita_notifications import notify_orbita_client
        target = request.POST.get("target", "")
        title = (request.POST.get("title") or "").strip()
        message_body = (request.POST.get("message") or "").strip()
        redirect_url = reverse("orbita_admin_notifications") + "?tab=enviar"
        if not title:
            messages.error(request, "El título es obligatorio.")
            return redirect(redirect_url)
        if target == "all":
            clients = ATSClient.objects.all()
            count = 0
            for client in clients:
                notify_orbita_client(
                    client, ATSNotification.TYPE_ADMIN, title,
                    message=message_body, request=request,
                )
                count += 1
            messages.success(request, f"Notificación enviada a {count} clientes.")
        else:
            client = get_object_or_404(ATSClient, pk=target)
            notify_orbita_client(
                client, ATSNotification.TYPE_ADMIN, title,
                message=message_body, request=request,
            )
            messages.success(request, f"Notificación enviada a {client.company_name}.")
        return redirect(redirect_url)


class ATSAdminNotificationsView(StaffRequiredMixin, View):
    """Página completa de notificaciones y solicitudes para el admin."""
    template_name = "orbita/admin/notificaciones.html"
    paginate_by = 30

    def get(self, request):
        tab = request.GET.get("tab", "solicitudes")
        pending = list(
            PlanChangeRequest.objects.filter(status=PlanChangeRequest.STATUS_PENDING)
            .select_related("client", "client__user")
            .order_by("-created_at")
        )
        history = PlanChangeRequest.objects.filter(
            status=PlanChangeRequest.STATUS_DONE
        ).select_related("client", "client__user").order_by("-created_at")
        history_paginator = Paginator(history, self.paginate_by)
        history_page = history_paginator.get_page(request.GET.get("hp", 1))
        all_clients = ATSClient.objects.order_by("company_name").values("id", "company_name")
        return render(request, self.template_name, {
            "orbita_page": "admin_notificaciones",
            "orbita_client": None,
            "tab": tab,
            "pending_requests": pending,
            "history_page": history_page,
            "all_clients_for_notify": list(all_clients),
        })


class ATSAdminMarkRequestDoneView(StaffRequiredMixin, View):
    """POST: marcar solicitud de cambio de plan como atendida."""
    http_method_names = ["post"]

    def post(self, request):
        req_id = request.POST.get("request_id")
        if req_id:
            PlanChangeRequest.objects.filter(pk=req_id).update(
                status=PlanChangeRequest.STATUS_DONE
            )
            messages.success(request, "Solicitud marcada como atendida.")
        return redirect("orbita_admin_notifications")


class ATSStaffAccountView(StaffRequiredMixin, View):
    """Mi cuenta para el administrador (staff): datos de usuario y enlace a cambiar contraseña."""
    template_name = "orbita/admin/mi_cuenta.html"

    def get(self, request):
        if request.user.is_staff and not getattr(request.user, "ats_client", None):
            return render(request, self.template_name, {
                "orbita_page": "mi_cuenta_staff",
                "orbita_client": None,
                "orbita_header_title": "Mi cuenta",
            })
        return redirect("orbita_admin_dashboard")


class ATSPasswordChangeView(StaffRequiredMixin, AuthPasswordChangeView):
    """Cambio de contraseña para el administrador ATS (mismo layout)."""
    template_name = "orbita/admin/password_change_form.html"
    success_url = reverse_lazy("orbita_staff_account")
    login_url = reverse_lazy("orbita_plataforma")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["orbita_page"] = "mi_cuenta_staff"
        context["orbita_client"] = None
        context["orbita_header_title"] = "Cambiar contraseña"
        return context


# --- Recuperar contraseña (cliente y staff) ---

class ATSPasswordResetView(AuthPasswordResetView):
    template_name = "orbita/auth/password_reset_form.html"
    success_url = reverse_lazy("orbita_password_reset_done")
    email_template_name = "orbita/auth/email/password_reset_email.html"
    subject_template_name = "orbita/auth/email/password_reset_subject.txt"


class ATSPasswordResetDoneView(AuthPasswordResetDoneView):
    template_name = "orbita/auth/password_reset_done.html"


class ATSPasswordResetConfirmView(AuthPasswordResetConfirmView):
    template_name = "orbita/auth/password_reset_confirm.html"
    success_url = reverse_lazy("orbita_password_reset_complete")
    post_reset_login = False


class ATSPasswordResetCompleteView(AuthPasswordResetCompleteView):
    template_name = "orbita/auth/password_reset_complete.html"














